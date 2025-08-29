import datetime
import os
import re
from tempfile import TemporaryDirectory
import threading
from tkinter.font import Font
import traceback
import zipfile
from dotenv import load_dotenv
from flask import Blueprint, Response, app, jsonify, logging, render_template, render_template_string, request
import mysql.connector
import openpyxl
import pdfkit
import requests
import urllib
from config import get_db_connection
from flask import jsonify, request
from flask import current_app
import pyodbc





timetable_bp = Blueprint('timetable', __name__, url_prefix='/api/timetable')



@timetable_bp.route('/filter', methods=['GET'])
def get_timetable_filtered():
    connection = get_db_connection()
    cursor = connection.cursor()  # ✅ no as_dict
    try:
        years = request.args.get('years')
        semester = request.args.get('semester')
        teacher_id = request.args.get('teacher_id', '').strip()
        shift_name = request.args.get('shift_name', '').strip()
        generation = request.args.get('generation', '').strip()
        group_student = request.args.get('group_student', '').strip()
        major_id = request.args.get('major_id', '').strip()
        room_id = request.args.get('room_id', '').strip()

        if not years or not semester:
            return jsonify({'error': 'សូមបញ្ជាក់ years និង semester'}), 400

        is_teacher_query = bool(teacher_id and teacher_id.lower() != "null")  # ✅ handle "null"
        is_student_query = all([generation, group_student, major_id])
        is_room_query = bool(room_id)

        query = """
            SELECT 
                t.id, t.note, t.study_sessions_id, t.group_student,
                t.batch, t.generation, t.major_id, t.teacher_id, t.subject_id, t.room_id, 
                t.years, t.semester,
                CONVERT(VARCHAR(8), session_time_start, 108) AS session_time_start,
                CONVERT(VARCHAR(8), session_time_end, 108) AS session_time_end,
                ss.shift_name AS study_shift_name,  
                ss.sessions_day AS study_session_day,
                m.name AS major_name,
                ts.teacher_id AS teacher_id, 
                ts.subject_id AS subject_id,
                r.room_number AS room_number,
                sub.name AS subject_name
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN teacher_subjects ts ON t.teacher_id = ts.teacher_id AND t.subject_id = ts.subject_id
            LEFT JOIN Rooms r ON t.room_id = r.id
            LEFT JOIN Subjects sub ON t.subject_id = sub.id
            WHERE t.years = ? AND t.semester = ?
        """
        params = [years, semester]

        if is_room_query and not is_teacher_query and not is_student_query:
            query += " AND t.room_id = ?"
            params.append(room_id)
        elif is_teacher_query and not is_student_query:
            query += " AND t.teacher_id = ?"
            params.append(teacher_id)
        elif not is_teacher_query and is_student_query:
            query += " AND (? = '' OR ss.shift_name = ?) AND t.generation = ? AND t.group_student = ? AND t.major_id = ?"
            params.extend([shift_name, shift_name, generation, group_student, major_id])
        else:
            return jsonify({'error': 'សូមបញ្ជាក់ teacher_id ឬ all student-group parameters ឬ room_id'}), 400

        if is_room_query:
            query += " AND t.room_id = ?"
            params.append(room_id)

        cursor.execute(query, tuple(params))

        # ✅ fetch as dicts
        columns = [col[0] for col in cursor.description]
        rows = cursor.fetchall()
        timetable_entries = [dict(zip(columns, row)) for row in rows]

        # Unique group info
        unique_group_info_set = {
            (e['batch'], e['generation'], e['major_id'], e['group_student'], e['major_name'], e['subject_id'], e['subject_name'], e['study_shift_name'])
            for e in timetable_entries
        }

        unique_group_info_list = [
            {
                'batch': batch,
                'generation': generation,
                'major_id': major_id,
                'group_student': group_student,
                'major_name': major_name,
                'subject_id': subject_id,
                'subject_name': subject_name,
                'study_shift_name': study_shift_name
            }
            for (batch, generation, major_id, group_student, major_name, subject_id, subject_name, study_shift_name)
            in unique_group_info_set
        ]

        return jsonify({
            'timetable_entries': timetable_entries,
            'info': unique_group_info_list
        })

    except Exception as e:
        current_app.logger.error(f"Error fetching timetable data: {e}")
        return jsonify({'error': 'មានបញ្ហាក្នុងការទាញយកទិន្នន័យ'}), 500

    finally:
        cursor.close()
        connection.close()



@timetable_bp.route('/options', methods=['GET'])
def get_timetable_options():
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        queries = {
            "years": "SELECT DISTINCT years FROM Timetable ORDER BY years ASC",
            "semester": "SELECT DISTINCT semester FROM Timetable ORDER BY semester ASC",
            "group_student": "SELECT DISTINCT group_student FROM Timetable ORDER BY group_student ASC",
            "batch": "SELECT DISTINCT batch FROM Timetable ORDER BY batch ASC",
            "generation": "SELECT DISTINCT generation FROM Timetable ORDER BY generation ASC",
            "shift_name": """
                SELECT DISTINCT ss.shift_name 
                FROM Timetable t 
                LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id 
                ORDER BY ss.shift_name ASC
            """,
            "teacher": """
                SELECT DISTINCT t.teacher_id, te.name AS teacher_name
                FROM Timetable t
                LEFT JOIN teachers te ON t.teacher_id = te.id
                ORDER BY t.teacher_id ASC
            """,
            "major": """
                SELECT DISTINCT m.id AS major_id, m.name AS major_name
                FROM Timetable t
                LEFT JOIN majors m ON t.major_id = m.id
                ORDER BY m.name ASC
            """,
            "room": """
                SELECT DISTINCT r.id AS room_id, r.room_number 
                FROM Timetable t
                LEFT JOIN Rooms r ON t.room_id = r.id
                ORDER BY r.room_number ASC
            """
        }

        result = {}
        for key, query in queries.items():
            cursor.execute(query)
            columns = [col[0] for col in cursor.description]  # column names
            rows = cursor.fetchall()
            result[key] = [dict(zip(columns, row)) for row in rows]

        return jsonify(result)

    except Exception as e:
        current_app.logger.error(f"Error fetching timetable options: {e}")
        return jsonify({'error': 'មានបញ្ហាក្នុងការទាញយកទិន្នន័យ'}), 500
    finally:
        cursor.close()
        connection.close()

        
        
# not yet

@timetable_bp.route('/tree', methods=['GET'])
def get_timetable_tree():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        query = """
            SELECT 
                t.years,
                t.semester,
                t.generation,
                f.id AS faculty_id,
                f.name AS faculty_name,
                d.id AS department_id,
                d.name AS department_name,
                m.id AS major_id,
                m.name AS major_name,
                t.group_student
            FROM Timetable t
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN Departments d ON m.Departments_id = d.id
            LEFT JOIN Faculties f ON d.Faculties_id = f.id
            ORDER BY t.years, t.semester, t.generation, f.name, d.name, m.name, t.group_student
        """

        cursor.execute(query)
        data = cursor.fetchall()

        # Build a tree structure
        tree = {}

        for row in data:
            years = row["years"]
            semester = row["semester"]
            generation = row["generation"]
            faculty_id = row["faculty_id"]
            faculty_name = row["faculty_name"]
            department_id = row["department_id"]
            department_name = row["department_name"]
            major_id = row["major_id"]
            major_name = row["major_name"]
            group_student = row["group_student"]

            if years not in tree:
                tree[years] = {}

            if semester not in tree[years]:
                tree[years][semester] = {}

            if generation not in tree[years][semester]:
                tree[years][semester][generation] = {}

            if faculty_id not in tree[years][semester][generation]:
                tree[years][semester][generation][faculty_id] = {
                    "faculty_name": faculty_name,
                    "departments": {}
                }

            if department_id not in tree[years][semester][generation][faculty_id]["departments"]:
                tree[years][semester][generation][faculty_id]["departments"][department_id] = {
                    "department_name": department_name,
                    "majors": {}
                }

            if major_id not in tree[years][semester][generation][faculty_id]["departments"][department_id]["majors"]:
                tree[years][semester][generation][faculty_id]["departments"][department_id]["majors"][major_id] = {
                    "major_name": major_name,
                    "group_students": []
                }

            # Append unique group_student values
            if group_student not in tree[years][semester][generation][faculty_id]["departments"][department_id]["majors"][major_id]["group_students"]:
                tree[years][semester][generation][faculty_id]["departments"][department_id]["majors"][major_id]["group_students"].append(group_student)

        return jsonify(tree)

    except Exception as e:
        app.logger.error(f"Error fetching timetable hierarchy: {e}")
        return jsonify({'error': 'An error occurred while fetching timetable hierarchy'}), 500
    finally:
        cursor.close()
        connection.close()


from collections import defaultdict


@timetable_bp.route('auto', methods=['GET'])
def get_timetable_auto():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch timetable data with detailed error logging
        cursor.execute("""
            SELECT 
                t.id, 
                t.note, 
                t.study_sessions_id, 
                t.group_student, 
                t.batch, 
                t.generation, 
                t.major_id, 
                t.teacher_id, 
                t.subject_id, 
                t.room_id,
                t.years,
                t.semester,
                TIME_FORMAT(session_time_start, '%H:%i:%s') AS session_time_start,
                TIME_FORMAT(session_time_end, '%H:%i:%s') AS session_time_end,
                ss.shift_name AS study_shift_name,  
                ss.sessions_day AS study_session_day,
                m.name AS major_name,
                ts.teacher_id AS teacher_id, 
                ts.subject_id AS subject_id,
                r.room_number AS room_number,
                sub.name AS subject_name
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN teacher_subjects ts ON t.teacher_id = ts.teacher_id AND t.subject_id = ts.subject_id
            LEFT JOIN Rooms r ON t.room_id = r.id
            LEFT JOIN Subjects sub ON t.subject_id = sub.id
        """)
        timetable_entries = cursor.fetchall()

        # Group data by years > semester > major > generation > group_student
        grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list)))))

        for entry in timetable_entries:
            # Group by years > semester > major > generation > group_student
            grouped_data[entry['years']][entry['semester']][entry['major_id']][entry['generation']][entry['group_student']].append(entry)

        # Convert the grouped data to a JSON-compatible format
        # Nested dictionary iteration to make it JSON-compatible
        final_grouped_data = {}
        for year, year_data in grouped_data.items():
            final_grouped_data[year] = {}
            for semester, semester_data in year_data.items():
                final_grouped_data[year][semester] = {}
                for major_id, major_data in semester_data.items():
                    final_grouped_data[year][semester][major_id] = {}
                    for generation, generation_data in major_data.items():
                        final_grouped_data[year][semester][major_id][generation] = {}
                        for group_student, entries in generation_data.items():
                            final_grouped_data[year][semester][major_id][generation][group_student] = entries

        return jsonify({'grouped_timetable': final_grouped_data})

    except Exception as e:
        # Log the exception details
        app.logger.error(f"Error fetching timetable data: {e}")
        return jsonify({'error': 'An error occurred while fetching timetable data'}), 500
    finally:
        cursor.close()
        connection.close()








      
import pandas as pd
import zipfile
from io import BytesIO
from flask import send_file, jsonify




# @timetable_bp.route('/downloade', methods=['GET'])
# def get_timetable_autoexcel3():
#     connection = get_db_connection()
#     cursor = connection.cursor()

#     try:
#         # Fetch timetable data
#         cursor.execute("""
#             SELECT 
#                 t.id, 
#                 t.note, 
#                 t.study_sessions_id, 
#                 t.group_student, 
#                 t.batch, 
#                 t.generation, 
#                 t.major_id,
#                 m.name AS major_name, 
#                 t.teacher_id, 
#                 t.subject_id, 
#                 t.room_id,
#                 t.years,
#                 t.semester,
#                 TIME_FORMAT(session_time_start, '%H:%i:%s') AS session_time_start,
#                 TIME_FORMAT(session_time_end, '%H:%i:%s') AS session_time_end,
#                 ss.shift_name AS study_shift_name,  
#                 ss.sessions_day AS study_session_day,
#                 r.room_number AS room_number,
#                 sub.name AS subject_name,
#                 te.name AS teacher_name
#             FROM Timetable t
#             LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
#             LEFT JOIN Majors m ON t.major_id = m.id
#             LEFT JOIN Rooms r ON t.room_id = r.id
#             LEFT JOIN Subjects sub ON t.subject_id = sub.id
#             LEFT JOIN Teachers te ON t.teacher_id = te.id
#         """)
#         timetable_entries = cursor.fetchall()

#         if not timetable_entries:
#             return jsonify({'message': 'No timetable data found'}), 404

#         # Group data by years > semester > major > generation > group_student
#         grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list)  ))))
        
#         for entry in timetable_entries:
#             grouped_data[entry['years']][entry['semester']][entry['major_id']][entry['generation']][entry['group_student']].append(entry)

#         # Prepare an in-memory ZIP file
#         zip_buffer = BytesIO()
#         with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            
#             # Loop through groups to create separate Excel files
#             for year, year_data in grouped_data.items():
#                 for semester, semester_data in year_data.items():
#                     for major_id, major_data in semester_data.items():
#                         for generation, generation_data in major_data.items():
#                             for group_student, entries in generation_data.items():
                                
                                
            

#                                 rows = []
#                                 previous_values = {
#                                     'year': None,
#                                     'semester': None,
#                                     'major': None,
#                                     'generation': None,
#                                     'group_student': None,
#                                     'study_shift': None,
#                                     'sessions_day': None
#                                 }
#                                 for entry in entries:
#                                     year_value = year if year != previous_values['year'] else ''
#                                     semester_value = semester if semester != previous_values['semester'] else ''
#                                     major_value = entry['major_name'] if entry['major_name'] != previous_values['major'] else ''
#                                     generation_value = generation if generation != previous_values['generation'] else ''
#                                     group_student_value = group_student if group_student != previous_values['group_student'] else ''
#                                     study_shift_value = entry['study_shift_name'] if entry['study_shift_name'] != previous_values['study_shift'] else ''
#                                     sessions_day_value = entry['study_session_day'] if entry['study_session_day'] != previous_values['sessions_day'] else ''

                                    
                                
#                                     rows.append({
                                       
                     
#                                         'Session Start': entry['session_time_start'],
#                                         'Session End': entry['session_time_end'],
#                                         'Sessions Day': sessions_day_value,
#                                         'Subject Name': entry['subject_name'],
#                                         'Teacher Name': entry['teacher_name'],
#                                         'Room Number': entry['room_number'],
    
#                                     })

#                                     previous_values.update({
#                                         'study_shift': entry['study_shift_name'],
#                                         'sessions_day': entry['study_session_day'],
#                                         'semester': semester,
#                                         'major': entry['major_name'],
#                                         'generation': generation,
#                                         'group_student': group_student,
                                    
#                                     })
                                    
                   
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                

                                
#                                                                 # Initialize HTML content with CSS styling
#                                 html_content = """
#                                 <html>
#                                 <html lang="en">
#                                 <head>
#                                     <meta charset="UTF-8">
#                                     <meta name="viewport" content="width=device-width, initial-scale=1.0">
                             
#                                 </head>
                                
#                                        <style>
#                                         body {
#                                                             font-family: Moul;
#                                                             font-size: xx-small;
#                                                         }
#                                                         table {
#                                                             width: 100%;
#                                                             border-collapse: collapse;
#                                                             font-family: Battambang;
#                                                         }
#                                                         th, td {
#                                                             border: 1px solid black;
#                                                             padding: 10px;
#                                                             text-align: center;
#                                                         }
#                                                         .header {
#                                                             text-align: center;
#                                                             margin-bottom: 20px;
#                                             }
#                                     </style>
                                
                                
                                
                                
#                                 <div class="header" style=" display: flex;">


#                                     <div  style="width: 20%;">
                                    
#                                         <div>
#                                             <br>
#                                             <img src="https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQEan2MdKi7N57D2ZeaZ7Y3iUauNq99Tdopcw&s"
#                                                 alt="University Logo" style="width:100px;">
#                                         </div>
                                    
#                                         <div>
#                                             <p>សាកលវិទ្យាល័យជាតិមានជ័យ
#                                                 <br>
#                                                 ការិយាល័យសិក្សា
#                                             </p>
#                                         </div>
#                                     </div>

                                
                                                                                
#                                   <br>             
                                                    
                                                    
                            
#                                 """
                                
#                                 html_content +="""
#                                           <div style="background-color: rgba(0, 255, 255, 0); width: 100%; transform: translateX(-10%);">
#                                                 ព្រះរាជាណាចក្រកម្ពុជា
#                                                 <br>
#                                                 ជាតិ សាសនា ព្រះមហាក្សត្រ

                                            
#                                                 <br>
#                                                 <img width="100px" src="https://puyvongheng.github.io/img/Screenshot%202024-11-13%20105603.png" alt="">
#                                                 <br>
#                                                 <br>
#                                                 <br>
#                                                 <br>
                                                
                   
                                                
                                                
                                                
                                     
#                                 """
                                
                                
#                             #    html_content += f"<p>      ជំនាន់ទី {generation}, ក្រុមទី {group_student}    ឆ្នាំ  {year} កាលវិភាគប្រចាំឆមាសទី {semester}, វេនសិក្សា ជំនាញ  {major_value}, "

                              
                                
#                                 html_content += f"""<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">​​   ​</span>
#                                 """
#                                 html_content += f"  ឆ្នាំសិក្សា   {year} - {year +1} កាលវិភាគប្រចាំឆមាសទី {semester}  "
                                
#                                 html_content += f"""<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">​​ <br>  ​</span>
#                                 """
                                
#                                 html_content += f"""<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">​​ ចាប់ផ្តើមពីថ្ងៃ....កើត ខែ.... ឆ្នាំ.... .... ត្រូវនឹងថ្ងៃទី.... ខែ....ឆ្នាំ.... វេនសិក្សា {entry['study_shift_name']} ​</span>
#                                 """
                                
#                                 html_content += f"""<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">​​ <br>  ​</span>
#                                 """

#                                 html_content += f"  ជំនាញ{entry['major_name']}  ជំនាន់ទី {generation}  </p> "


#                                 html_content += """
                                
        
                                               
                
            
                                                
#                                             </div>
                                            
#                                     </div>
#                                     <!--  vvvv-->
#                                     <br>
#                                 """
                                
#                                 # Khmer labels for days
#                                 day_labels = {
#                                     "monday": "ថ្ងៃច័ន្ទ/Monday",
#                                     "tuesday": "ថ្ងៃអង្គារ/Tuesday",
#                                     "wednesday": "ថ្ងៃពុធ/Wednesday",
#                                     "thursday": "ថ្ងៃព្រហស្បតិ៍/Thursday",
#                                     "friday": "ថ្ងៃសុក្រ/Friday",
#                                     "saturday": "ថ្ងៃសៅរ៍/Saturday",
#                                     "sunday": "ថ្ងៃអាទិត្យ/Sunday"
#                                 }

#                                 weekdays = ["monday", "tuesday", "wednesday", "thursday", "friday"]
#                                 weekends = ["saturday", "sunday"]

#                                 # Dictionary to group sessions by time
#                                 grouped_sessions = {}

#                                 # Process entries and group by session time
#                                 for entry in entries:
#                                     session_time_start = entry["session_time_start"]  # Keep original format (e.g., "09:00:00")
#                                     session_time_end = entry["session_time_end"]
#                                     session_time = f"{session_time_start} - {session_time_end}"
#                                     day = entry['study_session_day'].lower()  # Convert to lowercase for consistency
                                    
#                                     subject_details = (
#                                         f"{entry['subject_name']}<br>"
#                                         f"{entry['study_shift_name']}<br>"
#                                         f"{entry['teacher_name']}<br>"
#                                         f"{entry['room_number']}"
#                                     )

#                                     # Initialize session time entry if not exists
#                                     if session_time not in grouped_sessions:
#                                         grouped_sessions[session_time] = {day: "" for day in weekdays + weekends}

#                                     # Append subject details to the respective day
#                                     if day in grouped_sessions[session_time]:
#                                         grouped_sessions[session_time][day] += subject_details + "<br>"

#                                 # **Sort sessions by session_time_start (String Comparison Works for HH:MM:SS Format)**
#                                 sorted_sessions = sorted(grouped_sessions.keys(), key=lambda t: t.split(" - ")[0])

#                                 # Determine if weekdays and weekends should be displayed
#                                 show_weekdays = any(grouped_sessions[time][day].strip() for time in grouped_sessions for day in weekdays)
#                                 show_weekends = any(grouped_sessions[time][day].strip() for time in grouped_sessions for day in weekends)

#                                 # Start HTML content
#                                 html_content += """                               
#                                 <body style="margin: 10px;">
#                                     <table>
#                                         <thead id="timetable">
#                                             <tr>
#                                                 <th>ម៉ោងសិក្សា</th>"""  # Time column

#                                 # Add only required weekday headers
#                                 if show_weekdays:
#                                     for day in weekdays:
#                                         html_content += f"<th>{day_labels[day]}</th>"

#                                 # Add only required weekend headers
#                                 if show_weekends:
#                                     for day in weekends:
#                                         html_content += f"<th>{day_labels[day]}</th>"

#                                 html_content += """</tr></thead><tbody>"""

#                                 # Generate table rows dynamically in sorted order
#                                 for session_time in sorted_sessions:
#                                     html_content += f"<tr><td>{session_time}</td>"
                                    
#                                     if show_weekdays:
#                                         for day in weekdays:
#                                             html_content += f"<td>{grouped_sessions[session_time][day]}</td>"
                                    
#                                     if show_weekends:
#                                         for day in weekends:
#                                             html_content += f"<td>{grouped_sessions[session_time][day]}</td>"
                                    
#                                     html_content += "</tr>"

#                                 # Close table and body
#                                 html_content += """
#                                         </tbody>
#                                     </table>
#                                 </body>
#                                 """
                                
#                                 html_content += """
                                
#                                             <div style="display: flex; font-family: Battambang;">
#                                                     <div class="footer" style="text-align: center; margin-top: 20px; width: 50%; ">
#                                                         <p style=" font-family: Moul;">
#                                                         <br>
#                                                         <br>
#                                                         <br><br><br>
#                                                         <span style="font-family: Battambang;">
#                                                         បានឃើញ និងឯកភាព
#                                                         </span>
#                                                         <br>
#                                                         ជ.សាកលវិទ្យាធិការ
#                                                         <br>
#                                                         សាកលវិទ្យាធិការរង
#                                                         <br>
#                                                         </p>
#                                                     </div>

#                                                 <div class="footer" style="text-align: center; margin-top: 20px; width: 50%;">
#                                                     <p>ថ្ងៃទី<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">.......</span> ខែ<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">.......</span> ឆ្ន<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">.......</span> <span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">.......</span></p>
#                                                     <p >បន្ទាយមានជ័យ ថ្ងៃទី<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">.......</span> ខែ<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">.......</span> ឆ្នាំ<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">.......</span></p>
#                                                     <p>ប្រធានការិយាល័យសិក្សា</p>
#                                                 </div>
#                                             </div>

#                                 </html>
#                                 """

                     
                                                            
            
#                                 directory_path = f"html/ឆ្នាំ{year}/ ឆមាស{semester}/ជំានាន{generation}/ជំនាញ{entry['major_name']}"





#                                 os.makedirs(directory_path, exist_ok=True)
#                                 filename_html = f"{directory_path}/{entry['study_shift_name']} {year}__ ឆមាស{semester}__{entry['major_name']} {major_id}{major_value}__ ជំានាន{generation}__ ក្រុម{group_student}.html"

                                
    
                                
#                                # filename_html = f" html/ វេនសិក្សា​ {entry['study_shift_name']} ឆ្នាំ​ {year}__ ឆមាស{semester}__{entry['major_name']} ជំាញ {major_id}{major_value}__ ជំានាន{generation}__ ក្រុម{group_student}.html"
#                                 zip_file.writestr(filename_html, html_content)
                                
                 
                                
                    
                    
#                                                     # Insert custom rows (header and filename)
#                                 # Add Big Title, Subtitle, and Header to the rows list
#     # Insert Big Title and Subtitle into rows
#                                 rows.insert(0, {
                                    
#                                     'year': 'ឆ្នាំ {} ឆមាស {}   ជំាញ {} ជំានាន {} ក្រុម {} {}'.format(year, semester , major_value, generation, group_student ,study_shift_value),
#                                     'Session End': '',
#                                     'Sessions Day': '',
#                                     'Subject Name': '', 
#                                     'Teacher Name': '', 
#                                     'Room Number': '', 
                            
#                                 })

#                                 rows.insert(1, {
#                                     'Session Start': '',
#                                     'Session End': '',
#                                     'Sessions Day':  '',
#                                     'Subject Name': '', 
#                                     'Teacher Name': '', 
#                                     'Room Number': '', 
#                                 })

#                                 # Insert Column Headers
#                                 rows.insert(2, {
                            
#                                     'Session Start': 'Session Start', 
#                                     'Session End': 'Session End',
#                                     'Sessions Day': 'Sessions Day', 
#                                     'Subject Name': 'Subject Name', 
#                                     'Teacher Name': 'Teacher Name', 
#                                     'Room Number': 'Room Number', 
                                
#                                 })
#                                 # Convert to DataFrame
                                
                                
#                                 filename = f"excel/ ឆ្នាំ{year}__ ឆមាស{semester}__ ជំាញ{major_value}__ ជំានាន{generation}__ ក្រុម{group_student}.xlsx"
                              
#                                 df = pd.DataFrame(rows)
                                
                                
                           
                          
                                
                            
        
#                                 # Create in-memory Excel file
#                                 excel_buffer = BytesIO()
#                                 with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
#                                     df.to_excel(writer, index=False, sheet_name="Timetable")
                                    
                                    
#                                       # Get the workbook and worksheet
#                                     worksheet = writer.sheets["Timetable"]
                                    
                         
#                                     # Adjust column width based on the length of data in each column
#                                     for col_num in range(2, worksheet.max_column + 1):  # Start from column 2 (B)
#                                         column = worksheet.cell(row=1, column=col_num).column_letter  # Get the column letter (e.g. 'B', 'C', ...)
#                                         max_length = 0
#                                         for cell in worksheet[column]:
#                                             try:
#                                                 if len(str(cell.value)) > max_length:
#                                                     max_length = len(cell.value)
#                                             except:
#                                                 pass
#                                         adjusted_width = max_length + 2  # Add a little padding
#                                         worksheet.column_dimensions[column].width = adjusted_width




#                                     # Apply styles (e.g., bold headers)
#                                     for cell in worksheet["1:1"]:
#                                         cell.font = openpyxl.styles.Font(bold=True)
                                        
#                                     worksheet.delete_rows(1)
                                    
                                
                                        
                                        
                                    
#                                         # Apply style to row 2 (Subtitle Row) specifically
#                                     for col_num in range(1, worksheet.max_column + 1):
#                                         cell = worksheet.cell(row=3, column=col_num)
#                                         cell.alignment = Alignment(horizontal='center', vertical='center')
#                                         cell.font = Font(bold=True)


                                          
#                                     fill_light_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
#                                     for col_num in range(1, worksheet.max_column + 1):
#                                         cell = worksheet.cell(row=3, column=col_num)
#                                         cell.fill = fill_light_blue 
                                         
#                                     for col_num in range(1, worksheet.max_column + 1):
#                                         cell = worksheet.cell(row=1, column=col_num)
#                                         cell.alignment = Alignment(horizontal='center', vertical='center')                                                                                                                           
#                                         worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=worksheet.max_column)
                                        
                                        
                                        
                                    
                                                                      
   

#                                 excel_buffer.seek(0)

#                                 # Generate unique filename
                           
#                                 # Add to ZIP file
#                                 zip_file.writestr(filename, excel_buffer.getvalue())

#         zip_buffer.seek(0)
        
        
  
#      #   telegram_sent = send_telegram_file1(zip_buffer, "Timetable_Files.zip")

    
  
    
      
#         # Return ZIP file as response
#         return send_file(

#             zip_buffer,
#           # as_attachment=True,
#             download_name="Timetable_Files.zip",
#             mimetype="application/zip",
#         )
           

#     except Exception as e:
#         app.logger.error(f"Error fetching timetable data: {str(e)}")
#         return jsonify({'error': f'An error occurred while fetching timetable data: {str(e)}'}), 500

#     finally:
#         cursor.close()
#         connection.close()
     
     
     
# dowloadn zip file sqlserver
@timetable_bp.route('/downloade', methods=['GET'])
def get_timetable_autoexcel3():
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor()  # Remove dictionary=True

    def row_to_dict(cursor, row):
        """Convert a row tuple to a dictionary using column names from cursor.description."""
        return {cursor.description[i][0]: value for i, value in enumerate(row)} if row else None

    try:
        # Fetch timetable data with SQL Server-compatible time formatting
        cursor.execute("""
            SELECT 
                t.id, 
                t.note, 
                t.study_sessions_id, 
                t.group_student, 
                t.batch, 
                t.generation, 
                t.major_id,
                m.name AS major_name, 
                t.teacher_id, 
                t.subject_id, 
                t.room_id,
                t.years,
                t.semester,
                CONVERT(VARCHAR(8), ss.session_time_start, 108) AS session_time_start,
                CONVERT(VARCHAR(8), ss.session_time_end, 108) AS session_time_end,
                ss.shift_name AS study_shift_name,  
                ss.sessions_day AS study_session_day,
                r.room_number AS room_number,
                sub.name AS subject_name,
                te.name AS teacher_name
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN Rooms r ON t.room_id = r.id
            LEFT JOIN Subjects sub ON t.subject_id = sub.id
            LEFT JOIN Teachers te ON t.teacher_id = te.id
        """)
        timetable_entries = [row_to_dict(cursor, row) for row in cursor.fetchall()]

        if not timetable_entries:
            return jsonify({'message': 'No timetable data found'}), 404

        # Group data by years > semester > major > generation > group_student
        grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list)))))
        
        for entry in timetable_entries:
            grouped_data[entry['years']][entry['semester']][entry['major_id']][entry['generation']][entry['group_student']].append(entry)

        # Prepare an in-memory ZIP file
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Loop through groups to create separate Excel and HTML files
            for year, year_data in grouped_data.items():
                for semester, semester_data in year_data.items():
                    for major_id, major_data in semester_data.items():
                        for generation, generation_data in major_data.items():
                            for group_student, entries in generation_data.items():
                                major_value = entries[0]['major_name']  # Get major_name from first entry
                                study_shift_value = entries[0]['study_shift_name']  # Get shift_name from first entry

                                rows = []
                                previous_values = {
                                    'year': None,
                                    'semester': None,
                                    'major': None,
                                    'generation': None,
                                    'group_student': None,
                                    'study_shift': None,
                                    'sessions_day': None
                                }
                                for entry in entries:
                                    year_value = year if year != previous_values['year'] else ''
                                    semester_value = semester if semester != previous_values['semester'] else ''
                                    major_value = entry['major_name'] if entry['major_name'] != previous_values['major'] else ''
                                    generation_value = generation if generation != previous_values['generation'] else ''
                                    group_student_value = group_student if group_student != previous_values['group_student'] else ''
                                    study_shift_value = entry['study_shift_name'] if entry['study_shift_name'] != previous_values['study_shift'] else ''
                                    sessions_day_value = entry['study_session_day'] if entry['study_session_day'] != previous_values['sessions_day'] else ''

                                    rows.append({
                                        'Session Start': entry['session_time_start'],
                                        'Session End': entry['session_time_end'],
                                        'Sessions Day': sessions_day_value,
                                        'Subject Name': entry['subject_name'],
                                        'Teacher Name': entry['teacher_name'],
                                        'Room Number': entry['room_number'],
                                    })

                                    previous_values.update({
                                        'study_shift': entry['study_shift_name'],
                                        'sessions_day': entry['study_session_day'],
                                        'semester': semester,
                                        'major': entry['major_name'],
                                        'generation': generation,
                                        'group_student': group_student,
                                    })

                                # Initialize HTML content with CSS styling
                                html_content = """
                                <html lang="en">
                                <head>
                                    <meta charset="UTF-8">
                                    <meta name="viewport" content="width=device-width, initial-scale=1.0">
                                </head>
                                <style>
                                    body {
                                        font-family: Moul;
                                        font-size: xx-small;
                                    }
                                    table {
                                        width: 100%;
                                        border-collapse: collapse;
                                        font-family: Battambang;
                                    }
                                    th, td {
                                        border: 1px solid black;
                                        padding: 10px;
                                        text-align: center;
                                    }
                                    .header {
                                        text-align: center;
                                        margin-bottom: 20px;
                                    }
                                </style>
                                <div class="header" style="display: flex;">
                                    <div style="width: 20%;">
                                        <div>
                                            <br>
                                            <img src="https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQEan2MdKi7N57D2ZeaZ7Y3iUauNq99Tdopcw&s" alt="University Logo" style="width:100px;">
                                        </div>
                                        <div>
                                            <p>សាកលវិទ្យាល័យជាតិមានជ័យ
                                                <br>
                                                ការិយាល័យសិក្សា
                                            </p>
                                        </div>
                                    </div>
                                    <div style="background-color: rgba(0, 255, 255, 0); width: 100%; transform: translateX(-10%);">
                                        ព្រះរាជាណាចក្រកម្ពុជា
                                        <br>
                                        ជាតិ សាសនា ព្រះមហាក្សត្រ
                                        <br>
                                        <img width="100px" src="https://puyvongheng.github.io/img/Screenshot%202024-11-13%20105603.png" alt="">
                                        <br><br><br><br>
                                    </div>
                                </div>
                                <br>
                                """
                                html_content += f"""
                                    <p>ឆ្នាំសិក្សា {year} - {year + 1} កាលវិភាគប្រចាំឆមាសទី {semester}</p>
                                    <span contenteditable="true" style="border-bottom: 1px rgb(0, 0, 0);"></span>
                                    <p>ចាប់ផ្តើមពីថ្ងៃ....កើត ខែ.... ឆ្នាំ.... .... ត្រូវនឹងថ្ងៃទី.... ខែ....ឆ្នាំ.... វេនសិក្សា {study_shift_value}</p>
                                    <span contenteditable="true" style="border-bottom: 1px rgb(0, 0, 0);"></span>
                                    <p>ជំនាញ {major_value} ជំនាន់ទី {generation}</p>
                                    <br>
                                """

                                # Khmer labels for days
                                day_labels = {
                                    "monday": "ថ្ងៃច័ន្ទ/Monday",
                                    "tuesday": "ថ្ងៃអង្គារ/Tuesday",
                                    "wednesday": "ថ្ងៃពុធ/Wednesday",
                                    "thursday": "ថ្ងៃព្រហស្បតិ៍/Thursday",
                                    "friday": "ថ្ងៃសុក្រ/Friday",
                                    "saturday": "ថ្ងៃសៅរ៍/Saturday",
                                    "sunday": "ថ្ងៃអាទិត្យ/Sunday"
                                }

                                weekdays = ["monday", "tuesday", "wednesday", "thursday", "friday"]
                                weekends = ["saturday", "sunday"]

                                # Dictionary to group sessions by time
                                grouped_sessions = {}
                                for entry in entries:
                                    session_time_start = entry["session_time_start"]
                                    session_time_end = entry["session_time_end"]
                                    session_time = f"{session_time_start} - {session_time_end}"
                                    day = entry['study_session_day'].lower()
                                    
                                    subject_details = (
                                        f"{entry['subject_name']}<br>"
                                        f"{entry['study_shift_name']}<br>"
                                        f"{entry['teacher_name']}<br>"
                                        f"{entry['room_number']}"
                                    )

                                    if session_time not in grouped_sessions:
                                        grouped_sessions[session_time] = {day: "" for day in weekdays + weekends}
                                    if day in grouped_sessions[session_time]:
                                        grouped_sessions[session_time][day] += subject_details + "<br>"

                                sorted_sessions = sorted(grouped_sessions.keys(), key=lambda t: t.split(" - ")[0])

                                show_weekdays = any(grouped_sessions[time][day].strip() for time in grouped_sessions for day in weekdays)
                                show_weekends = any(grouped_sessions[time][day].strip() for time in grouped_sessions for day in weekends)

                                html_content += """
                                <body style="margin: 10px;">
                                    <table>
                                        <thead id="timetable">
                                            <tr>
                                                <th>ម៉ោងសិក្សា</th>"""
                                if show_weekdays:
                                    for day in weekdays:
                                        html_content += f"<th>{day_labels[day]}</th>"
                                if show_weekends:
                                    for day in weekends:
                                        html_content += f"<th>{day_labels[day]}</th>"
                                html_content += "</tr></thead><tbody>"

                                for session_time in sorted_sessions:
                                    html_content += f"<tr><td>{session_time}</td>"
                                    if show_weekdays:
                                        for day in weekdays:
                                            html_content += f"<td>{grouped_sessions[session_time][day]}</td>"
                                    if show_weekends:
                                        for day in weekends:
                                            html_content += f"<td>{grouped_sessions[session_time][day]}</td>"
                                    html_content += "</tr>"

                                html_content += """
                                        </tbody>
                                    </table>
                                    <div style="display: flex; font-family: Battambang;">
                                        <div class="footer" style="text-align: center; margin-top: 20px; width: 50%;">
                                            <p style="font-family: Moul;">
                                                <br><br><br><br><br>
                                                <span style="font-family: Battambang;">
                                                    បានឃើញ និងឯកភាព
                                                </span>
                                                <br>
                                                ជ.សាកលវិទ្យាធិការ
                                                <br>
                                                សាកលវិទ្យាធិការរង
                                                <br>
                                            </p>
                                        </div>
                                        <div class="footer" style="text-align: center; margin-top: 20px; width: 50%;">
                                            <p>ថ្ងៃទី<span contenteditable="true" style="border-bottom: 1px rgb(0, 0, 0);">.......</span> ខែ<span contenteditable="true" style="border-bottom: 1px rgb(0, 0, 0);">.......</span> ឆ្នាំ<span contenteditable="true" style="border-bottom: 1px rgb(0, 0, 0);">.......</span></p>
                                            <p>បន្ទាយមានជ័យ ថ្ងៃទី<span contenteditable="true" style="border-bottom: 1px rgb(0, 0, 0);">.......</span> ខែ<span contenteditable="true" style="border-bottom: 1px rgb(0, 0, 0);">.......</span> ឆ្នាំ<span contenteditable="true" style="border-bottom: 1px rgb(0, 0, 0);">.......</span></p>
                                            <p>ប្រធានការិយាល័យសិក្សា</p>
                                        </div>
                                    </div>
                                </body>
                                </html>
                                """

                                directory_path = f"html/ឆ្នាំ{year}/ឆមាស{semester}/ជំនាន់{generation}/ជំនាញ{major_value}"
                                os.makedirs(directory_path, exist_ok=True)
                                filename_html = f"{directory_path}/{study_shift_value} {year}__ឆមាស{semester}__{major_value}__ជំនាន់{generation}__ក្រុម{group_student}.html"
                                zip_file.writestr(filename_html, html_content)

                                # Insert custom rows for Excel
                                rows.insert(0, {
                                    'year': f'ឆ្នាំ {year} ឆមាស {semester} ជំនាញ {major_value} ជំនាន់ {generation} ក្រុម {group_student} {study_shift_value}',
                                    'Session Start': '',
                                    'Session End': '',
                                    'Sessions Day': '',
                                    'Subject Name': '',
                                    'Teacher Name': '',
                                    'Room Number': '',
                                })
                                rows.insert(1, {
                                    'Session Start': '',
                                    'Session End': '',
                                    'Sessions Day': '',
                                    'Subject Name': '',
                                    'Teacher Name': '',
                                    'Room Number': '',
                                })
                                rows.insert(2, {
                                    'Session Start': 'Session Start',
                                    'Session End': 'Session End',
                                    'Sessions Day': 'Sessions Day',
                                    'Subject Name': 'Subject Name',
                                    'Teacher Name': 'Teacher Name',
                                    'Room Number': 'Room Number',
                                })

                                # Create Excel file
                                filename = f"excel/ឆ្នាំ{year}__ឆមាស{semester}__ជំនាញ{major_value}__ជំនាន់{generation}__ក្រុម{group_student}.xlsx"
                                df = pd.DataFrame(rows)
                                excel_buffer = BytesIO()
                                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                    df.to_excel(writer, index=False, sheet_name="Timetable")
                                    worksheet = writer.sheets["Timetable"]

                                    # Adjust column widths
                                    for col_num in range(2, worksheet.max_column + 1):
                                        column = worksheet.cell(row=1, column=col_num).column_letter
                                        max_length = 0
                                        for cell in worksheet[column]:
                                            try:
                                                if len(str(cell.value)) > max_length:
                                                    max_length = len(cell.value)
                                            except:
                                                pass
                                        adjusted_width = max_length + 2
                                        worksheet.column_dimensions[column].width = adjusted_width

                                    # Apply styles
                                    for cell in worksheet["1:1"]:
                                        cell.font = Font(bold=True)
                                    worksheet.delete_rows(1)

                                    for col_num in range(1, worksheet.max_column + 1):
                                        cell = worksheet.cell(row=3, column=col_num)
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                        cell.font = Font(bold=True)
                                        cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

                                    for col_num in range(1, worksheet.max_column + 1):
                                        cell = worksheet.cell(row=1, column=col_num)
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                    worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=worksheet.max_column)

                                excel_buffer.seek(0)
                                zip_file.writestr(filename, excel_buffer.getvalue())

        zip_buffer.seek(0)

        # Return ZIP file as response
        return send_file(
            zip_buffer,
            download_name="Timetable_Files.zip",
            mimetype="application/zip",
        )

    except pyodbc.Error as e:
        app.logger.error(f"Error fetching timetable data: {str(e)}")
        return jsonify({'error': f'SQL Server Error: {str(e)}'}), 500
    finally:
        cursor.close()
        connection.close()
     
        
from openpyxl.styles import PatternFill
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

import pandas as pd
from openpyxl.styles import Alignment

















# mysql
# @timetable_bp.route('', methods=['GET'])
# def get_timetable():
#     connection = get_db_connection()
#     cursor = connection.cursor(dictionary=True)

#     try:
#         # Fetch timetable data with detailed error logging
#         cursor.execute("""
#             SELECT 
#                 t.id, 
#                 t.note, 
#                 t.study_sessions_id, 
#                 t.group_student, 
#                 t.batch, 
#                 t.generation, 
#                 t.major_id, 
#                 t.teacher_id, 
#                 t.subject_id, 
#                 t.room_id,
#                         t.years,
#                         t.semester,
                       
#                 TIME_FORMAT(session_time_start, '%H:%i:%s') AS session_time_start,
#                 TIME_FORMAT(session_time_end, '%H:%i:%s') AS session_time_end,
#                 ss.shift_name AS study_shift_name,  
#                 ss.sessions_day AS study_session_day,
#                 m.name AS major_name,
#                 ts.teacher_id AS teacher_id, 
#                 ts.subject_id AS subject_id,
#                 r.room_number AS room_number,
#                 sub.name AS subject_name
                 
#             FROM Timetable t
#             LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
#             LEFT JOIN Majors m ON t.major_id = m.id
#             LEFT JOIN teacher_subjects ts ON t.teacher_id = ts.teacher_id AND t.subject_id = ts.subject_id
#             LEFT JOIN Rooms r ON t.room_id = r.id
#             LEFT JOIN Subjects sub ON t.subject_id = sub.id
         
#         """)
#         timetable_entries = cursor.fetchall()

#         return jsonify({'timetable_entries': timetable_entries})

#     except Exception as e:
#         # Log the exception details
#         app.logger.error(f"Error fetching timetable data: {e}")
#         return jsonify({'error': 'An error occurred while fetching timetable data'}), 500
#     finally:
#         cursor.close()
#         connection.close()




# sqlserver 
@timetable_bp.route('', methods=['GET'])
def get_timetable():
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor()  # Remove dictionary=True

    def row_to_dict(cursor, row):
        """Convert a row tuple to a dictionary using column names from cursor.description."""
        return {cursor.description[i][0]: value for i, value in enumerate(row)} if row else None

    try:
        # Fetch timetable data with SQL Server-compatible time formatting
        cursor.execute("""
            SELECT 
                t.id, 
                t.note, 
                t.study_sessions_id, 
                t.group_student, 
                t.batch, 
                t.generation, 
                t.major_id, 
                t.teacher_id, 
                t.subject_id, 
                t.room_id,
                t.years,
                t.semester,
                CONVERT(VARCHAR(8), ss.session_time_start, 108) AS session_time_start,
                CONVERT(VARCHAR(8), ss.session_time_end, 108) AS session_time_end,
                ss.shift_name AS study_shift_name,  
                ss.sessions_day AS study_session_day,
                m.name AS major_name,
                r.room_number AS room_number,
                sub.name AS subject_name,
                te.name AS teacher_name
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN Rooms r ON t.room_id = r.id
            LEFT JOIN Subjects sub ON t.subject_id = sub.id
            LEFT JOIN Teachers te ON t.teacher_id = te.id
        """)
        timetable_entries = [row_to_dict(cursor, row) for row in cursor.fetchall()]

        return jsonify({'timetable_entries': timetable_entries})

    except pyodbc.Error as e:
        app.logger.error(f"SQL Server Error fetching timetable data: {e}")
        return jsonify({'error': f'SQL Server Error: {e}'}), 500
    finally:
        cursor.close()
        connection.close()
      
      
      
      

# 🔍 Check  my sql 
# @timetable_bp.route('/checkconflicts', methods=['POST'])
# def check_timetable_conflicts():
#     connection = get_db_connection()
#     cursor = connection.cursor(dictionary=True)

#     data = request.json
#     required_fields = [ 'years', 'semester']
#     missing_fields = [field for field in required_fields if not data.get(field)]
#     if missing_fields:
#         return jsonify({'error': f'Missing required fields: {missing_fields}'}), 400

#     # Extract required parameters
#     study_sessions_id = data.get('study_sessions_id')
#     years = data.get('years')
#     semester = data.get('semester')
#     teacher_id = data.get('teacher_id')
#     room_id = data.get('room_id')
#     subject_id = data.get('subject_id')
#     major_id = data.get('major_id')
#     generation = data.get('generation')
#     group_student = data.get('group_student')

#     try:
#         conflicts = []
        
        
        
        
        

#         # cheak generation គ្រូ បង្រៀ​ន
#         cursor.execute("""
#             SELECT 
#             t.id, 
#             t.group_student, 
#             t.generation, 
#             m.name AS major_name
#             FROM Timetable t
#             JOIN majors m ON t.major_id = m.id
#             WHERE t.teacher_id = %s 
#             AND t.study_sessions_id = %s 
#             AND t.years = %s 
#             AND t.semester = %s 
#             AND t.generation != %s
#         """, (teacher_id, study_sessions_id, years, semester, generation))

#         # Fetch all conflicting rows
#         conflicting_data_rows = cursor.fetchall()
#         conflict_messages = []
#         if conflicting_data_rows:
#             for row in conflicting_data_rows:
#                 timetable_generation = row['generation']
#                 major_name = row['major_name']
#                 group =row['group_student']
#                 conflict_messages.append(f'- ជំនា {timetable_generation} ជំនាញ ({major_name} {group}) ')
#             conflict_line = ', '.join(conflict_messages)
#             conflicts.append({
#                 'teacher': f'​🚫​ គ្រូម្នាក់មិនអាចបង្រៀនជំនាន់ផ្សេងគ្នា គ្រូជាប់បង្រៀន {conflict_line}'
#             })
    

#         cursor.fetchall()  # This is just a safety check, though fetchall() has already been called earlier








#         # Count the total number of timetables for the given year, semester, major, and generation
#         cursor.execute("""
#             SELECT COUNT(*) AS total_timetables 
#             FROM Timetable 
#             WHERE years = %s 
#             AND semester = %s 
#             AND major_id = %s 
#             AND generation = %s
#             AND group_student = %s
#         """, (years, semester, major_id, generation ,group_student))

#         total_timetables = cursor.fetchone()['total_timetables']
#         if total_timetables :
#             conflicts.append({
#                 'total': f'  {total_timetables} sessions  mix 15',
#             })
#         else:
#             conflicts.append({
#                 'total': f'  មិនទាន់មានកាវិភាគសោះ',
#             })
            
#         cursor.fetchall()
        
        
        
        
 
            
            
            
            
#         cursor.execute("""
#         SELECT 
#             t.id, 
#             t.group_student, 
#             t.generation, 
#             m.name AS major_name,
#             r.room_number AS room,
#             s.name AS subject_name  
#         FROM Timetable t
#         JOIN majors m ON t.major_id = m.id
#         JOIN rooms r ON t.room_id = r.id
#         JOIN subjects s ON t.subject_id = s.id  
#         WHERE t.teacher_id = %s 
#         AND t.years = %s 
#         AND t.semester = %s 
#         AND t.study_sessions_id = %s
#     """, (teacher_id, years, semester, study_sessions_id))

#         room_subject_exists = cursor.fetchall()

#         if room_subject_exists:
#             conflict_messages = []
#             seen_combinations = {}  # Store subject_name + room as key and group details as value

#             for row in room_subject_exists:
#                 timetable_generation = row['generation']
#                 major_name = row['major_name']
#                 group = row['group_student']
#                 room = row['room']
#                 subject_name = row['subject_name']

#                 key = (subject_name, room)  # Unique key for subject and room

#                 if key not in seen_combinations:  # If it's a new subject-room combination
#                     seen_combinations[key] = []  # Initialize list to hold group details

#                 # Add the new group to the list for this subject-room combination
#                 seen_combinations[key].append(f' + ({major_name} {group} - Generation {timetable_generation})')

#             # Build the conflict message based on collected details
#             for (subject_name, room), groups in seen_combinations.items():
#                 group_details = ', '.join(groups)  # Combine all group details into one string
#                 conflict_messages.append(f'- {subject_name} បន្ទប់: {room} - {group_details}')

#             # Join all conflict messages into a single line, separated by commas
#             conflict_line = ', '.join(conflict_messages)

#             # Append the formatted message to conflicts
#             conflicts.append({
#                 'room_subject': f'​ គ្រូបានដាក់អោយបង្រៀនហើយ 🤝 អាចបន្ថែមសិស្សក្រុមផ្សេងបាន: {conflict_line}'
#             })
#         cursor.fetchall()  # Ensure we fetch all rows here to avoid unread result












#         # Check for room conflicts
#         cursor.execute("""
#             SELECT id, room_id, study_sessions_id, years, semester 
#             FROM Timetable 
#             WHERE room_id = %s AND study_sessions_id = %s AND years = %s AND semester = %s
#         """, (room_id, study_sessions_id, years, semester))
#         room_conflict = cursor.fetchone()
#         if room_conflict:
#             conflicts.append({
#                 'room': '🚫 បន្ទប់នេះមិនទំនេរ!',
#             })
#         cursor.fetchall()  # Ensure we fetch all rows here to avoid unread result

#         # Check for teacher conflicts
#         cursor.execute("""
#             SELECT id, teacher_id, study_sessions_id, years, semester 
#             FROM Timetable 
#             WHERE teacher_id = %s AND study_sessions_id = %s AND years = %s AND semester = %s
#         """, (teacher_id, study_sessions_id, years, semester))
#         teacher_conflict = cursor.fetchone()
#         if teacher_conflict:
#             conflicts.append({
#                 'teacher': '🚫 គ្រូរមិនទំនេរ!',
#             })
#         cursor.fetchall()  # Ensure we fetch all rows here to avoid unread result




            
     
        

#         cursor.execute("""
#             SELECT COUNT(*) FROM Timetable 
#             WHERE years = %s AND semester = %s AND major_id = %s AND generation = %s AND group_student = %s
#         """, (years, semester, major_id, generation, group_student))

#         count = cursor.fetchone()['COUNT(*)']
#         if count >= 15:
#             conflicts.append({
#                 'messagesessions': '🚫 ពុំអាចបញ្ចូលការបង្រៀននេះទេ ព្រោះមានចំនួនសិស្សច្រើនជាង 15 នាក់!',
#             })






#         # Check for group student conflicts
#         cursor.execute("""
#             SELECT * FROM Timetable 
#             WHERE 
#             group_student = %s 
#             AND study_sessions_id = %s 
#             AND years = %s 
#             AND semester = %s 
#             AND major_id = %s 
#             AND generation = %s
#         """, (group_student, study_sessions_id, years, semester, major_id, generation))
#         group_student_conflict = cursor.fetchone()
#         if group_student_conflict:
#             conflicts.append({
#                 'messagesessions': '🚫 មានម៉ោងសិក្សារហើយ!'
#             })
#         cursor.fetchall()  
        
        
        
        


#         cursor.execute("""
#             SELECT ss.shift_name
#             FROM study_sessions ss
#             JOIN Timetable t ON ss.id = t.study_sessions_id
#             WHERE t.group_student = %s 
#             AND t.years = %s 
#             AND t.semester = %s 
#             AND t.major_id = %s 
#             AND t.generation = %s
#         """, (group_student, years, semester, major_id, generation))
#         existing_shifts = cursor.fetchall()
#         if existing_shifts:
#             shift_names = [shift['shift_name'].lower() for shift in existing_shifts]  # convert to lowercase for case-insensitive comparison

#             if len(set(shift_names)) > 4:
#                 conflicts.append({
#                     'messagesessions': '🚫 សិស្សក្រុមនេះមិនអាចរៀននៅ shift_name ផ្សេងគ្នា! សូមជ្រើសរើស shift_name តែមួយដែលស្រប!'
#                 })
#             else:
#                 cursor.execute("""
#                     SELECT shift_name
#                     FROM study_sessions
#                     WHERE id = %s
#                 """, (study_sessions_id,))
#                 new_shift_entry = cursor.fetchone()
#                 if new_shift_entry:
#                     new_shift_name = new_shift_entry['shift_name'].lower()
#                     if new_shift_name == shift_names[0]:
#                         conflicting_shifts = "  និង ".join(set(shift_names))
                        
                     
#                         conflicts.append({
#                                 'messagesessions_g': f'✅ នៅវេន:  {conflicting_shifts}✅ '
#                         })
                        
                        
#                     else:
#                         conflicting_shifts = ", ".join(set(shift_names))
    
#                         if new_shift_name in shift_names:
#                             conflicts.append({
#                                 'messagesessions': f'✅ នៅវេន {conflicting_shifts}  គ្មានបញ្ហា! {new_shift_name}'
#                             })
#                         else:
#                             if len(set(shift_names)) > 1:
#                                 conflicts.append({
#                                     'messagesessions': f'🚫 នៅវេន {conflicting_shifts} និង {new_shift_name} មានការបញ្ហា!'
#                                 })
#                             else:
#                                 conflicts.append({
#                                     'messagesessions': f'⁉️ នៅវេន {conflicting_shifts}, ហើយចងនៅ នៅវេន {new_shift_name} ទៀតមេន?'
#                                 })
                    
   
  

#         # Check if the teacher is available at the study session
#         cursor.execute("""
#             SELECT * FROM teacher_teaching_time
#             WHERE teacher_id = %s AND study_sessions_id = %s
#         """, (teacher_id, study_sessions_id))
#         if not cursor.fetchone():
#             conflicts.append({'teacher_time': '🚫 គ្រូមិនអាចបង្រៀននៅម៉ោងនេះបានទេ'})
#         cursor.fetchall()  # Ensure we fetch all rows here to avoid unread result
        
        
        

#         # Return conflicts if found
#         if conflicts:
#             return jsonify({'conflicts': conflicts}), 200
#         else:
#             return jsonify({'message': 'No conflicts found'}), 200

#     except mysql.connector.Error as err:
#         # MySQL specific error logging
#         print(f"MySQL Error: {err}")
#         return jsonify({'error': f'MySQL Error: {err}'}), 500
#     except Exception as e:
#         # General error logging
#         print(f"An error occurred: {str(e)}")
#         return jsonify({'error': f'An error occurred: {str(e)}'}), 500
#     finally:
#         # Ensure all results are fetched before closing the cursor
#         try:
#             cursor.fetchall()  # Fetch any remaining results to avoid 'Unread result' error
#         except Exception as e:
#             print(f"Error while fetching remaining results: {str(e)}")
#         cursor.close()  # Close the cursor
#         connection.close()  # Close the connection


@timetable_bp.route('/checkconflicts', methods=['POST'])
def check_timetable_conflicts():
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor()  # Remove dictionary=True

    data = request.json
    required_fields = ['years', 'semester']
    missing_fields = [field for field in required_fields if not data.get(field)]
    if missing_fields:
        return jsonify({'error': f'Missing required fields: {missing_fields}'}), 400

    # Extract required parameters
    study_sessions_id = data.get('study_sessions_id')
    years = data.get('years')
    semester = data.get('semester')
    teacher_id = data.get('teacher_id')
    room_id = data.get('room_id')
    subject_id = data.get('subject_id')
    major_id = data.get('major_id')
    generation = data.get('generation')
    group_student = data.get('group_student')

    def row_to_dict(cursor, row):
        """Convert a row tuple to a dictionary using column names from cursor.description."""
        return {cursor.description[i][0]: value for i, value in enumerate(row)}

    try:
        conflicts = []

        # Check generation conflicts for teacher
        cursor.execute("""
            SELECT 
                t.id, 
                t.group_student, 
                t.generation, 
                m.name AS major_name
            FROM Timetable t
            JOIN Majors m ON t.major_id = m.id
            WHERE t.teacher_id = ? 
                AND t.study_sessions_id = ? 
                AND t.years = ? 
                AND t.semester = ? 
                AND t.generation != ?
        """, (teacher_id, study_sessions_id, years, semester, generation))

        # Fetch all conflicting rows
        conflicting_data_rows = [row_to_dict(cursor, row) for row in cursor.fetchall()]
        conflict_messages = []
        if conflicting_data_rows:
            for row in conflicting_data_rows:
                timetable_generation = row['generation']
                major_name = row['major_name']
                group = row['group_student']
                conflict_messages.append(f'- ជំនា {timetable_generation} ជំនាញ ({major_name} {group}) ')
            conflict_line = ', '.join(conflict_messages)
            conflicts.append({
                'teacher': f'🚫 គ្រូម្នាក់មិនអាចបង្រៀនជំនាន់ផ្សេងគ្នា គ្រូជាប់បង្រៀន {conflict_line}'
            })

        # Count the total number of timetables
        cursor.execute("""
            SELECT COUNT(*) AS total_timetables 
            FROM Timetable 
            WHERE years = ? 
                AND semester = ? 
                AND major_id = ? 
                AND generation = ? 
                AND group_student = ?
        """, (years, semester, major_id, generation, group_student))

        total_timetables = row_to_dict(cursor, cursor.fetchone())['total_timetables']
        if total_timetables:
            conflicts.append({
                'total': f'{total_timetables} sessions mix 15',
            })
        else:
            conflicts.append({
                'total': f'មិនទាន់មានកាវិភាគសោះ',
            })

        # Check room and subject conflicts for teacher
        cursor.execute("""
            SELECT 
                t.id, 
                t.group_student, 
                t.generation, 
                m.name AS major_name,
                r.room_number AS room,
                s.name AS subject_name  
            FROM Timetable t
            JOIN Majors m ON t.major_id = m.id
            JOIN Rooms r ON t.room_id = r.id
            JOIN Subjects s ON t.subject_id = s.id  
            WHERE t.teacher_id = ? 
                AND t.years = ? 
                AND t.semester = ? 
                AND t.study_sessions_id = ?
        """, (teacher_id, years, semester, study_sessions_id))

        room_subject_exists = [row_to_dict(cursor, row) for row in cursor.fetchall()]
        if room_subject_exists:
            conflict_messages = []
            seen_combinations = {}
            for row in room_subject_exists:
                timetable_generation = row['generation']
                major_name = row['major_name']
                group = row['group_student']
                room = row['room']
                subject_name = row['subject_name']
                key = (subject_name, room)
                if key not in seen_combinations:
                    seen_combinations[key] = []
                seen_combinations[key].append(f' + ({major_name} {group} - Generation {timetable_generation})')
            for (subject_name, room), groups in seen_combinations.items():
                group_details = ', '.join(groups)
                conflict_messages.append(f'- {subject_name} បន្ទប់: {room} - {group_details}')
            conflict_line = ', '.join(conflict_messages)
            conflicts.append({
                'room_subject': f'គ្រូបានដាក់អោយបង្រៀនហើយ 🤝 អាចបន្ថែមសិស្សក្រុមផ្សេងបាន: {conflict_line}'
            })

        # Check for room conflicts
        cursor.execute("""
            SELECT id, room_id, study_sessions_id, years, semester 
            FROM Timetable 
            WHERE room_id = ? AND study_sessions_id = ? AND years = ? AND semester = ?
        """, (room_id, study_sessions_id, years, semester))
        room_conflict = cursor.fetchone()
        if room_conflict:
            conflicts.append({
                'room': '🚫 បន្ទប់នេះមិនទំនេរ!',
            })

        # Check for teacher conflicts
        cursor.execute("""
            SELECT id, teacher_id, study_sessions_id, years, semester 
            FROM Timetable 
            WHERE teacher_id = ? AND study_sessions_id = ? AND years = ? AND semester = ?
        """, (teacher_id, study_sessions_id, years, semester))
        teacher_conflict = cursor.fetchone()
        if teacher_conflict:
            conflicts.append({
                'teacher': '🚫 គ្រូរមិនទំនេរ!',
            })

        # Check session count
        cursor.execute("""
            SELECT COUNT(*) AS count 
            FROM Timetable 
            WHERE years = ? AND semester = ? AND major_id = ? AND generation = ? AND group_student = ?
        """, (years, semester, major_id, generation, group_student))
        count = row_to_dict(cursor, cursor.fetchone())['count']
        if count >= 15:
            conflicts.append({
                'messagesessions': '🚫 ពុំអាចបញ្ចូលការបង្រៀននេះទេ ព្រោះមានចំនួនសិស្សច្រើនជាង 15 នាក់!',
            })

        # Check for group student conflicts
        cursor.execute("""
            SELECT * FROM Timetable 
            WHERE group_student = ? 
                AND study_sessions_id = ? 
                AND years = ? 
                AND semester = ? 
                AND major_id = ? 
                AND generation = ?
        """, (group_student, study_sessions_id, years, semester, major_id, generation))
        group_student_conflict = cursor.fetchone()
        if group_student_conflict:
            conflicts.append({
                'messagesessions': '🚫 មានម៉ោងសិក្សារហើយ!'
            })

        # Check shift consistency
        cursor.execute("""
            SELECT ss.shift_name
            FROM study_sessions ss
            JOIN Timetable t ON ss.id = t.study_sessions_id
            WHERE t.group_student = ? 
                AND t.years = ? 
                AND t.semester = ? 
                AND t.major_id = ? 
                AND t.generation = ?
        """, (group_student, years, semester, major_id, generation))
        existing_shifts = [row_to_dict(cursor, row) for row in cursor.fetchall()]
        if existing_shifts:
            shift_names = [shift['shift_name'].lower() for shift in existing_shifts]
            if len(set(shift_names)) > 4:
                conflicts.append({
                    'messagesessions': '🚫 សិស្សក្រុមនេះមិនអាចរៀននៅ shift_name ផ្សេងគ្នា! សូមជ្រើសរើស shift_name តែមួយដែលស្រប!'
                })
            else:
                cursor.execute("""
                    SELECT shift_name
                    FROM study_sessions
                    WHERE id = ?
                """, (study_sessions_id,))
                new_shift_entry = cursor.fetchone()
                if new_shift_entry:
                    new_shift_name = new_shift_entry[0].lower()
                    if new_shift_name == shift_names[0]:
                        conflicting_shifts = " និង ".join(set(shift_names))
                        conflicts.append({
                            'messagesessions_g': f'✅ នៅវេន: {conflicting_shifts}✅ '
                        })
                    else:
                        conflicting_shifts = ", ".join(set(shift_names))
                        if new_shift_name in shift_names:
                            conflicts.append({
                                'messagesessions': f'✅ នៅវេន {conflicting_shifts} គ្មានបញ្ហា! {new_shift_name}'
                            })
                        else:
                            if len(set(shift_names)) > 1:
                                conflicts.append({
                                    'messagesessions': f'🚫 នៅវេន {conflicting_shifts} និង {new_shift_name} មានការបញ្ហា!'
                                })
                            else:
                                conflicts.append({
                                    'messagesessions': f'⁉️ នៅវេន {conflicting_shifts}, ហើយចងនៅ នៅវេន {new_shift_name} ទៀតមេន?'
                                })

        # Check if the teacher is available at the study session
        cursor.execute("""
            SELECT * FROM teacher_teaching_time
            WHERE teacher_id = ? AND study_sessions_id = ?
        """, (teacher_id, study_sessions_id))
        if not cursor.fetchone():
            conflicts.append({'teacher_time': '🚫 គ្រូមិនអাঅាចបង្រៀននៅម៉ោងនេះបានទេ'})

        # Return conflicts if found
        if conflicts:
            return jsonify({'conflicts': conflicts}), 200
        else:
            return jsonify({'message': 'No conflicts found'}), 200

    except pyodbc.Error as err:
        print(f"SQL Server Error: {err}")
        return jsonify({'error': f'SQL Server Error: {err}'}), 500
    finally:
        cursor.close()
        connection.close()








# create timetable  mysql
# @timetable_bp.route('', methods=['POST'])
# def create_timetable():
#     """✏️ បង្កើតកាលវិភាគថ្មី"""
#     data = request.json
#     current_app.logger.info(f"Received data: {data}")
#     # 🛑 ពិនិត្យពាក្យចាំបាច់ 🛑
#     required_fields = ['study_sessions_id', 'group_student', 'batch', 'generation',
#                        'major_id', 'teacher_id', 'subject_id', 'room_id', 'years', 'semester']
    
    
#     # Define a dictionary to map English field names to Khmer labels
#     khmer_labels = {
#         'study_sessions_id': 'វគ្គសិក្សា',
#         'group_student': 'ក្រុមសិស្ស',
#         'batch': 'ឆ្នាំ',
#         'generation': 'ជំនាន់',
#         'major_id': 'ជំនាញ',
#         'teacher_id': 'គ្រូបង្រៀន',
#         'subject_id': 'មុខវិជ្ជា',
#         'room_id': 'បន្ទប់',
#         'years': 'ឆ្នាំសិក្សា',
#         'semester': 'ឆមាស'
#     }

#     # Find missing fields
#     missing_fields = [khmer_labels[field] for field in required_fields if not data.get(field)]


#    # missing_fields = [field for field in required_fields if not data.get(field)]
#     errors = []
#     if missing_fields:
#         errors.append({'error': f"❌ ភ្លេច បញ្ចូល  <br>  <br>{' <br>  '.join(missing_fields)}"}), 400

#     # 📌 ទាញយកព័ត៌មានពី Request
#     note = data.get('note', '').strip()
#     study_sessions_id = data.get('study_sessions_id')
#     group_student = data.get('group_student')
#     batch = data.get('batch')
#     generation = data.get('generation')
#     major_id = data.get('major_id')
#     teacher_id = data.get('teacher_id')
#     subject_id = data.get('subject_id')
#     room_id = data.get('room_id')
#     years = data.get('years')
#     semester = data.get('semester')

#     connection = get_db_connection()
#     cursor = connection.cursor()

#     try:
        


#         cursor.execute("""
#             SELECT ss.shift_name
#             FROM study_sessions ss
#             JOIN Timetable t ON ss.id = t.study_sessions_id
#             WHERE t.group_student = %s 
#             AND t.years = %s 
#             AND t.semester = %s 
#             AND t.major_id = %s 
#             AND t.generation = %s
#         """, (group_student, years, semester, major_id, generation))

#         existing_shifts = cursor.fetchall()
#         if existing_shifts:
#             shift_names = [shift[0].lower() for shift in existing_shifts]  # assuming shift_name is in the first column
#             if len(set(shift_names)) > 4:
#                 errors.append({'error': '🚫 Multiple shift names detected for this group! Please select one consistent shift.'})
#             else:
#                 cursor.execute("""
#                     SELECT shift_name
#                     FROM study_sessions
#                     WHERE id = %s
#                 """, (study_sessions_id,))
#                 new_shift_entry = cursor.fetchone()
#                 if new_shift_entry:
#                     new_shift_name = new_shift_entry[0].lower()
#                     if new_shift_name == shift_names[0]:
#                         pass
#                        # errors.append({'error': f'✅ The group is already assigned to {shift_names[0]}.'})
#                     else:
#                         errors.append({'error': '🚫 The shift name does not match with the study session shift.'})




        
        
#     # បន្ថែមការពិនិត្យថ្មី៖ គ្រូម្នាក់មិនអាចបង្រៀនជំនាន់ខុសគ្នា នៅក្នុង study_sessions_id, years, និង semester
#         cursor.execute("""
#             SELECT id FROM Timetable 
#             WHERE teacher_id = %s AND study_sessions_id = %s 
#             AND years = %s AND semester = %s AND generation != %s
#         """, (teacher_id, study_sessions_id, years, semester, generation))

#         if cursor.fetchone():
#             errors.append({'error': '🚫 គ្រូម្នាក់មិនអាចបង្រៀនជំនាន់ផ្សេងគ្នាក្នងម៉ោងសិក្សាដូចគ្នា!'})


        
        
#         cursor.execute("SELECT number_sessions FROM Teachers WHERE id = %s", (teacher_id,))
#         teacher = cursor.fetchone()
#         if not teacher:
#             errors.append({'error': '🚫 Teacher not found.'}), 400
#             return jsonify({'errors': errors}), 400

#         # Access number_sessions using index (0 for the first column)
#         number_sessions = teacher[0]  # teacher is a tuple, so we access the first element

#         # 🔍 Query to count the current timetable entries for the teacher
#         cursor.execute("""
#             SELECT COUNT(DISTINCT CONCAT(room_id, '-', study_sessions_id)) AS total_timetable
#             FROM Timetable
#             WHERE teacher_id = %s AND years = %s AND semester = %s
#         """, (teacher_id, years, semester))

#         result = cursor.fetchone()

#         # Access total_timetable using index (0 for the first column in the tuple)
#         total_timetable = result[0] if result else 0  # result is a tuple, so access the first element

#         # 🔍 Check if total timetable exceeds teacher's allowed sessions
#         if total_timetable >= number_sessions:
#             errors.append({
#                 'error': f"🚫 គ្រូបានឈានដល់ចំនួនអតិបរមានៃ sessions ({number_sessions})ហើយសម្រាប់ឆមាសនេះ"
#             })
#             return jsonify({'errors': errors}), 400
        
        
        
        
        
    
        
#         # 🔍 ពិនិត្យមើលថាបន្ទប់ត្រូវបានប្រើរួចឬអត់  ROOM
#         cursor.execute("""
#             SELECT t.id, ss.id AS study_sessions_id, t.years, t.semester, 
#                 m.name AS major_name, sub.name AS subject_name, te.name AS teacher_name
#             FROM Timetable t
#             LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
#             LEFT JOIN Majors m ON t.major_id = m.id
#             LEFT JOIN Subjects sub ON t.subject_id = sub.id
#             LEFT JOIN Teachers te ON t.teacher_id = te.id
#             WHERE t.room_id = %s AND t.study_sessions_id = %s AND t.years = %s AND t.semester = %s
#             AND NOT (t.teacher_id = %s AND t.subject_id = %s)
#         """, (room_id, study_sessions_id, years, semester, teacher_id, subject_id))
#         existing_room = cursor.fetchone()
#         if existing_room:
#             errors.append({
#                 'error': f'🚫 <strong>បន្ទប់មិនទំនេរ</strong> <br/>'
#                         f'<strong>Timetable ID:</strong> {existing_room[0]} '
#                         f'<strong>វគ្គសិក្សា:</strong> {existing_room[1]} '
#                         f'<strong>ឆ្នាំសិក្សា:</strong> {existing_room[2]} '
#                         f'<strong>ឆមាស:</strong> {existing_room[3]} <br/>'
#                         f'<strong>ជំនាញ:</strong> {existing_room[4]} <br/>'
#                         f'<strong>មុខវិជ្ជា:</strong> {existing_room[5]} <br/>'
#                         f'<strong>គ្រូបង្រៀន:</strong> {existing_room[6]} <br/>​ <hr>'
#             })









 








#         # 🔍 ពិនិត្យមើលថាគ្រូស្ថិតនៅក្នុងបន្ទប់តែមួយក្នុង study_sessions_id, years, semester
#         cursor.execute("""
#             SELECT t.id, t.room_id, t.study_sessions_id, t.years, t.semester
#             FROM Timetable t
#             WHERE t.teacher_id = %s 
#                 AND t.study_sessions_id = %s 
#                 AND t.years = %s 
#                 AND t.semester = %s 
#                 AND t.room_id <> %s
#         """, (teacher_id, study_sessions_id, years, semester, room_id))
#         existing_teacher_room = cursor.fetchone()
#         if existing_teacher_room:
#             errors.append({
#                 'error': f'🚫 <strong>គ្រូបជាប់ង្រៀនហើយ ប្រហែលជានៅបន្ទប់ផ្សេង</strong> <br/>'
#                         f'👨‍🏫 <strong>សម្រាប់sessionsនេះ</strong> <br/>'
                  
#             })
            
            
            

            
            
            
            
 





#         # 🔍 ពិនិត្យថាក្រុមសិស្សមានម៉ោងសិក្សារួចនៅឬអត់  stusen  AND generation = %s
#         cursor.execute("""
#                 SELECT * FROM Timetable 
#                 WHERE 
#                 group_student = %s AND 
#                 study_sessions_id = %s AND 
#                 years = %s AND 
#                 Semester = %s  AND 
#                 major_id = %s AND
#                 generation  = %s AND  
#                 batch = %s
#             """, (group_student, study_sessions_id, years, semester, major_id,generation,batch))
#         if cursor.fetchone():
#             errors.append({
#                     'error': '🚫 <strong>ក្រុមសិស្សរួចហើយ</strong> <br/>'
#                             '👨‍🎓 <strong>សម្រាប់វគ្គនេះ</strong> <br/> <hr>'
#             })
            
            
        
        
    

      
        







#         # 🔍 ពិនិត្យមើលថាគ្រូនេះអាចបង្រៀនមុខវិជ្ជានេះឬអត់ teacher
#         cursor.execute("""
#             SELECT * FROM teacher_subjects 
#             WHERE teacher_id = %s AND subject_id = %s
#         """, (teacher_id, subject_id))

#         if not cursor.fetchone():
#             errors.append({'error': '🚫 គ្រូមិនអាចបង្រៀនមុខវិជ្ជានេះបានទេ ❌'}), 400

#         # 🔍 ពិនិត្យមើលថាគ្រូមានសមត្ថភាពបង្រៀននៅម៉ោងនេះឬអត់ teacher
#         cursor.execute("""
#             SELECT * FROM teacher_teaching_time
#             WHERE teacher_id = %s AND study_sessions_id = %s
#         """, (teacher_id, study_sessions_id))

#         if not cursor.fetchone():
#             errors.append({'error': '🚫 គ្រូមិនអាចបង្រៀននៅម៉ោងនេះបានទេ ❌'}), 400
            
            
#         # 👉 ប្រសិនបើមានកំហុសណាមួយ
#         if errors:
#             return jsonify({'errors': errors}), 400

#         # ✅ បញ្ចូលទិន្នន័យកាលវិភាគថ្មី
#         cursor.execute("""
#             INSERT INTO Timetable (
#                 study_sessions_id, group_student, note, batch, generation, major_id,
#                 teacher_id, subject_id, room_id, years, Semester
#             )
#             VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
#         """, (study_sessions_id, group_student, note, batch, generation, major_id,
#               teacher_id, subject_id, room_id, years, semester))

#         connection.commit()
#         return jsonify({'message': '🎉 បានបង្កើតដោយជោគជ័យ! ✅'}), 201

#     except mysql.connector.Error as err:
#         connection.rollback()
#         current_app.logger.error(f"🚨 Error creating timetable: {err}")
#         errors.append({'error': f'❌ មានបញ្ហា៖ {err}'}), 500
#         return jsonify({'errors': errors}), 500
#     finally:
#         cursor.close()
#         connection.close()
# sqlserver



@timetable_bp.route('', methods=['POST'])
def create_timetable():
    data = request.json
    current_app.logger.info(f"Received data: {data}")

    required_fields = ['study_sessions_id', 'group_student', 'batch', 'generation',
                       'major_id', 'teacher_id', 'subject_id', 'room_id', 'years', 'semester']
    khmer_labels = {
        'study_sessions_id': 'វគ្គសិក្សា',
        'group_student': 'ក្រុមសិស្ស',
        'batch': 'ឆ្នាំ',
        'generation': 'ជំនាន់',
        'major_id': 'ជំនាញ',
        'teacher_id': 'គ្រូបង្រៀន',
        'subject_id': 'មុខវិជ្ជា',
        'room_id': 'បន្ទប់',
        'years': 'ឆ្នាំសិក្សា',
        'semester': 'ឆមាស'
    }

    missing_fields = [khmer_labels[field] for field in required_fields if not data.get(field)]
    errors = []
    if missing_fields:
        return jsonify({'errors': [{'error': f"❌ ភ្លេច បញ្ចូល <br><br>{' <br> '.join(missing_fields)}"}]}), 400

    note = data.get('note', '').strip()
    study_sessions_id = data.get('study_sessions_id')
    group_student = data.get('group_student')
    batch = data.get('batch')
    generation = data.get('generation')
    major_id = data.get('major_id')
    teacher_id = data.get('teacher_id')
    subject_id = data.get('subject_id')
    room_id = data.get('room_id')
    years = data.get('years')
    semester = data.get('semester')

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = conn.cursor()

    try:
        # Check Shift Name for Group Student
        cursor.execute("""
            SELECT ss.shift_name
            FROM study_sessions ss
            JOIN Timetable t ON ss.id = t.study_sessions_id
            WHERE t.group_student = ? AND t.years = ? AND t.semester = ? 
                AND t.major_id = ? AND t.generation = ?
        """, (group_student, years, semester, major_id, generation))
        existing_shifts = [row[0].lower() for row in cursor.fetchall()]
        if existing_shifts:
            shift_names = set(existing_shifts)
            if len(shift_names) > 1:
                errors.append({'error': '🚫 ក្រុមនេះមាន shift ខុសៗគ្នា។ ត្រូវប្រើ shift ដូចគ្នា!'})
            cursor.execute("SELECT shift_name FROM study_sessions WHERE id = ?", (study_sessions_id,))
            new_shift = cursor.fetchone()
            if new_shift and new_shift[0].lower() not in shift_names:
                errors.append({'error': '🚫 Shift name មិនត្រូវគ្នាជាមួយ study session'})

        # Check if teacher is teaching different generations in the same session
        cursor.execute("""
            SELECT id FROM Timetable
            WHERE teacher_id = ? AND study_sessions_id = ? AND years = ? AND semester = ?
                AND generation != ?
        """, (teacher_id, study_sessions_id, years, semester, generation))
        if cursor.fetchone():
            errors.append({'error': '🚫 គ្រូម្នាក់មិនអាចបង្រៀនជំនាន់ផ្សេងគ្នាក្នុងម៉ោងដូចគ្នា!'})

        # Check teacher's session limit
        cursor.execute("SELECT number_sessions FROM Teachers WHERE id = ?", (teacher_id,))
        teacher = cursor.fetchone()
        if not teacher:
            return jsonify({'errors': [{'error': '🚫 គ្រូមិនមានក្នុងប្រព័ន្ធ'}]}), 400
        number_sessions = teacher[0]
        cursor.execute("""
            SELECT COUNT(DISTINCT CONCAT(room_id, '-', study_sessions_id)) 
            FROM Timetable WHERE teacher_id = ? AND years = ? AND semester = ?
        """, (teacher_id, years, semester))
        total_timetable = cursor.fetchone()[0] or 0
        if total_timetable >= number_sessions:
            errors.append({'error': f'🚫 គ្រូបានឈានដល់អតិបរមា {number_sessions} sessions'})

        # Check room availability
        cursor.execute("""
            SELECT t.id, ss.id, t.years, t.semester, m.name, sub.name, te.name
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN Subjects sub ON t.subject_id = sub.id
            LEFT JOIN Teachers te ON t.teacher_id = te.id
            WHERE t.room_id = ? AND t.study_sessions_id = ? AND t.years = ? AND t.semester = ?
                AND NOT (t.teacher_id = ? AND t.subject_id = ?)
        """, (room_id, study_sessions_id, years, semester, teacher_id, subject_id))
        if cursor.fetchone():
            errors.append({'error': '🚫 បន្ទប់មិនទំនេរ'})

        # Check if teacher is in another room for this session
        cursor.execute("""
            SELECT id FROM Timetable
            WHERE teacher_id = ? AND study_sessions_id = ? AND years = ? AND semester = ?
                AND room_id <> ?
        """, (teacher_id, study_sessions_id, years, semester, room_id))
        if cursor.fetchone():
            errors.append({'error': '🚫 គ្រូនេះមានបង្រៀននៅបន្ទប់ផ្សេងរួចហើយក្នុង session នេះ'})

        # Check if group student is already scheduled
        cursor.execute("""
            SELECT id FROM Timetable
            WHERE group_student = ? AND study_sessions_id = ? AND years = ? AND semester = ?
                AND major_id = ? AND generation = ? AND batch = ?
        """, (group_student, study_sessions_id, years, semester, major_id, generation, batch))
        if cursor.fetchone():
            errors.append({'error': '🚫 ក្រុមសិស្សនេះមានវគ្គសិក្សារួចហើយ'})

        # Check if teacher can teach this subject (use composite key)
        cursor.execute("""
            SELECT teacher_id, subject_id FROM teacher_subjects WHERE teacher_id = ? AND subject_id = ?
        """, (teacher_id, subject_id))
        if not cursor.fetchone():
            errors.append({'error': '🚫 គ្រូមិនអាចបង្រៀនមុខវិជ្ជានេះបានទេ'})

        # Check if teacher can teach at this time (use composite key)
        cursor.execute("""
            SELECT teacher_id, study_sessions_id FROM teacher_teaching_time WHERE teacher_id = ? AND study_sessions_id = ?
        """, (teacher_id, study_sessions_id))
        if not cursor.fetchone():
            errors.append({'error': '🚫 គ្រូមិនអាចបង្រៀននៅម៉ោងនេះ'})

        # Return errors if any
        if errors:
            return jsonify({'errors': errors}), 400

        # Insert new timetable
        cursor.execute("""
            INSERT INTO Timetable (
                study_sessions_id, group_student, note, batch, generation, major_id,
                teacher_id, subject_id, room_id, years, semester
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (study_sessions_id, group_student, note, batch, generation, major_id,
              teacher_id, subject_id, room_id, years, semester))
        conn.commit()

        return jsonify({'message': '🎉 បានបង្កើតដោយជោគជ័យ! ✅'}), 201

    except pyodbc.Error as err:
        conn.rollback()
        current_app.logger.error(f"🚨 Error creating timetable: {err}")
        return jsonify({'errors': [{'error': f'❌ មានបញ្ហា៖ {err}'}]}), 500
    finally:
        cursor.close()
        conn.close()






@timetable_bp.route('swap-teachers-subjects', methods=['PUT'])
def swap_teachers_subjects():
    """🔄 Swap teachers and subjects across all timetables (SQL Server version)"""
    data = request.json
    timetable_id_1 = data.get('timetable_id_1')
    timetable_id_2 = data.get('timetable_id_2')

    if not timetable_id_1 or not timetable_id_2:
        return jsonify({'error': '❌ Timetable IDs are required'}), 400

    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # 🔍 Fetch timetable 1
        cursor.execute("SELECT teacher_id, subject_id, study_sessions_id FROM Timetable WHERE id = ?", (timetable_id_1,))
        timetable_1 = cursor.fetchone()

        # 🔍 Fetch timetable 2
        cursor.execute("SELECT teacher_id, subject_id, study_sessions_id FROM Timetable WHERE id = ?", (timetable_id_2,))
        timetable_2 = cursor.fetchone()

        if not timetable_1 or not timetable_2:
            return jsonify({'error': '❌ One or both timetable IDs are invalid'}), 401

        teacher_1, subject_1, session_1 = timetable_1
        teacher_2, subject_2, session_2 = timetable_2

        # --- Check if teachers are the same ---
        if teacher_1 == teacher_2:
            return jsonify({'error': '❌ Teachers are the same'}), 400

        # 🔍 Check if teacher 1 is already assigned to session 2
        cursor.execute("""
            SELECT t.name FROM Timetable tim
            JOIN Teachers t ON tim.teacher_id = t.id
            WHERE tim.teacher_id = ? AND tim.study_sessions_id = ?
        """, (teacher_1, session_2))
        if cursor.fetchone():
            return jsonify({'error': f'🚫 Teacher {teacher_1} is already assigned to session {session_2}'}), 400

        # 🔍 Check if teacher 2 is already assigned to session 1
        cursor.execute("""
            SELECT t.name FROM Timetable tim
            JOIN Teachers t ON tim.teacher_id = t.id
            WHERE tim.teacher_id = ? AND tim.study_sessions_id = ?
        """, (teacher_2, session_1))
        if cursor.fetchone():
            return jsonify({'error': f'🚫 Teacher {teacher_2} is already assigned to session {session_1}'}), 400
        
        
        
        
        
        # ✅ Check if teachers can teach at the session times​​​​​​​          new
        cursor.execute("SELECT 1 FROM teacher_teaching_time WHERE teacher_id = ? AND study_sessions_id = ?", (teacher_1, session_2))
        if not cursor.fetchone():
            return jsonify({'error': f'🚫 គ្រូរ {teacher_1} មិនអាចបង្រៀនម៉ោងនេះទេ {session_2}'}), 400
        cursor.execute("SELECT 1 FROM teacher_teaching_time WHERE teacher_id = ? AND study_sessions_id = ?", (teacher_2, session_1))
        if not cursor.fetchone():
            return jsonify({'error': f'🚫 គ្រូរ  {teacher_2} មិនអាចបង្រៀនម៉ោងនេះទេ  {session_1}'}), 400




        
        
        
        

        # 🔄 Swap teacher_id and subject_id for timetables 1
        cursor.execute("""
            SELECT id FROM Timetable WHERE teacher_id = ? AND subject_id = ? AND study_sessions_id = ?
        """, (teacher_1, subject_1, session_1))
        timetables_1 = cursor.fetchall()

        for t in timetables_1:
            t_id = t[0]
            cursor.execute("UPDATE Timetable SET teacher_id = ?, subject_id = ? WHERE id = ?", (teacher_2, subject_2, t_id))

        # 🔄 Swap teacher_id and subject_id for timetables 2
        cursor.execute("""
            SELECT id FROM Timetable WHERE teacher_id = ? AND subject_id = ? AND study_sessions_id = ?
        """, (teacher_2, subject_2, session_2))
        timetables_2 = cursor.fetchall()

        for t in timetables_2:
            t_id = t[0]
            cursor.execute("UPDATE Timetable SET teacher_id = ?, subject_id = ? WHERE id = ?", (teacher_1, subject_1, t_id))

        connection.commit()

        # 🔔 Send Telegram message
        response_message = f"✅ Swapped teachers {teacher_1} ↔ {teacher_2} and subjects {subject_1} ↔ {subject_2}"
        send_telegram_message(response_message)

        return jsonify({'message': response_message}), 200

    except pyodbc.Error as err:
        connection.rollback()
        return jsonify({'error': f'❌ Database error: {err}'}), 500

    finally:
        cursor.close()
        connection.close()





@timetable_bp.route('swap-rooms', methods=['PUT'])
def swap_rooms():
    """🔄 ប្តូរ Room សម្រាប់ Timetable ទាំងអស់ដែលមាន Years, Semester, Subject, និង Teacher ដូចគ្នា"""
    data = request.json
    timetable_id_1 = data.get('timetable_id_1')
    timetable_id_2 = data.get('timetable_id_2')

    if not timetable_id_1 or not timetable_id_2:
        return jsonify({'error': '❌ ត្រូវការបញ្ជាក់ Timetable ID ទាំងពីរ'}), 400

    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # 🔍 ទាញយក timetable 1
        cursor.execute("""
            SELECT room_id, years, semester, subject_id, teacher_id, study_sessions_id 
            FROM Timetable WHERE id = ?
        """, (timetable_id_1,))
        timetable_1 = cursor.fetchone()

        # 🔍 ទាញយក timetable 2
        cursor.execute("""
            SELECT room_id, years, semester, subject_id, teacher_id, study_sessions_id 
            FROM Timetable WHERE id = ?
        """, (timetable_id_2,))
        timetable_2 = cursor.fetchone()

        if not timetable_1 or not timetable_2:
            return jsonify({'error': '❌ Timetable ID មួយឬទាំងពីរមិនត្រឹមត្រូវ'}), 400

        # 👉 Unpack
        room_1, years_1, semester_1, subject_1, teacher_1, session_1 = timetable_1
        room_2, years_2, semester_2, subject_2, teacher_2, session_2 = timetable_2

        # 🔍 Room មិនដូចគ្នា
        if room_1 == room_2:
            return jsonify({'error': '⚠️ Room ទាំងពីរដូចគ្នា 🏫 មិនមានអ្វីត្រូវផ្លាស់ប្តូរ!'}), 400

        # 🔍 Optional: Check room conflict in other sessions
        # cursor.execute("""
        #     SELECT id FROM Timetable WHERE room_id = ? AND study_sessions_id = ?
        # """, (room_1, session_2))
        # if cursor.fetchone():
        #     return jsonify({'error': f'🚫 Room {room_1} ត្រូវបានប្រើរួចហើយក្នុង Study Session {session_2}'}), 400
        #
        # cursor.execute("""
        #     SELECT id FROM Timetable WHERE room_id = ? AND study_sessions_id = ?
        # """, (room_2, session_1))
        # if cursor.fetchone():
        #     return jsonify({'error': f'🚫 Room {room_2} ត្រូវបានប្រើរួចហើយក្នុង Study Session {session_1}'}), 400

        # 🔄 Update rooms ជា transaction
        cursor.execute("""
            UPDATE Timetable
            SET room_id = ?
            WHERE years = ? AND semester = ? AND subject_id = ? AND teacher_id = ? AND study_sessions_id = ?
        """, (room_2, years_1, semester_1, subject_1, teacher_1, session_1))

        cursor.execute("""
            UPDATE Timetable
            SET room_id = ?
            WHERE years = ? AND semester = ? AND subject_id = ? AND teacher_id = ? AND study_sessions_id = ?
        """, (room_1, years_2, semester_2, subject_2, teacher_2, session_2))

        connection.commit()
        # send_telegram_message("មានកាផ្លាស់ប្តូរ Room ✅")
        
        threading.Thread(target=send_msg_async, args=("មានកាផ្លាស់ប្តូរ Room ✅",)).start()

        return jsonify({'message': '✅ ប្តូរ Room បានជោគជ័យ!'}), 200

    except pyodbc.Error as err:
        connection.rollback()
        return jsonify({'error': f'❌ Database error: {err}'}), 500

    finally:
        cursor.close()
        connection.close()




def send_msg_async(msg):
    send_telegram_message(msg)





@timetable_bp.route('assign-room/<int:timetable_id>/<int:room_id>', methods=['PUT'])
def assign_room(timetable_id, room_id):
    """កំណត់បន្ទប់ទៅឱ្យធាតុ Timetable ជាក់លាក់មួយ ដោយពិនិត្យមើលជម្លោះ។"""
    
    connection = get_db_connection()
    cursor = connection.cursor()
    try:
        # ១. យក study_sessions_id, years, semester, teacher_id, subject_id សម្រាប់ timetable_id
        cursor.execute("""
            SELECT study_sessions_id, years, semester, teacher_id, subject_id, major_id 
            FROM Timetable WHERE id = ?
        """, (timetable_id,))
        result = cursor.fetchone()

        if not result:
            return jsonify({'error': '❌ រក​ឃើញ​លេខ​សម្គាល់ Timetable មិន​ឃើញ'}), 401

        study_sessions_id, years, semester, teacher_id, subject_id, major_id = result

        # ២. ពិនិត្យថា room_id មាននៅក្នុងតារាង Rooms
        cursor.execute("SELECT id FROM Rooms WHERE id = ?", (room_id,))
        if not cursor.fetchone():
            return jsonify({'error': f'❌ លេខសម្គាល់បន្ទប់ {room_id} មិនមាន'}), 402

        # ៣. ពិនិត្យជម្លោះបន្ទប់ (Room Conflict)
        cursor.execute("""
            SELECT t.id, ss.id AS study_sessions_id, t.years, t.semester, 
                   m.name AS major_name, sub.name AS subject_name, te.name AS teacher_name
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN Subjects sub ON t.subject_id = sub.id
            LEFT JOIN Teachers te ON t.teacher_id = te.id
            WHERE t.room_id = ? AND t.study_sessions_id = ? AND t.years = ? AND t.semester = ?
              AND NOT (t.teacher_id = ? AND t.subject_id = ?)
        """, (room_id, study_sessions_id, years, semester, teacher_id, subject_id))
        existing_room = cursor.fetchone()

        if existing_room:
            return jsonify({
                'error': f'🚫 <strong>បន្ទប់មិនទំនេរ</strong> <br/>'
                        f'<strong>Timetable ID:</strong> {existing_room[0]}<br/> '
                        f'<strong>វគ្គសិក្សា:</strong> {existing_room[1]}<br/> '
                        f'<strong>ឆ្នាំសិក្សា:</strong> {existing_room[2]}<br/> '
                        f'<strong>ឆមាស:</strong> {existing_room[3]} <br/>'
                        f'<strong>ជំនាញ:</strong> {existing_room[4]} <br/>'
                        f'<strong>មុខវិជ្ជា:</strong> {existing_room[5]} <br/>'
                        f'<strong>គ្រូបង្រៀន:</strong> {existing_room[6]} <br/>​ <hr>'
            }), 409

        # ៤. Update ALL matching timetable entries
        cursor.execute("""
            UPDATE Timetable
            SET room_id = ?
            WHERE years = ? AND semester = ? AND subject_id = ? AND teacher_id = ? AND study_sessions_id = ?
        """, (room_id, years, semester, subject_id, teacher_id, study_sessions_id))

        connection.commit()
        send_telegram_message("មានកាផ្លាស់ប្តូរ Room ✅")

        return jsonify({'message': f'✅ បន្ទប់ {room_id} បាន​កំណត់​ទៅ Timetable {timetable_id}'}), 200

    except pyodbc.Error as err:
        connection.rollback()
        return jsonify({'error': f'❌ កំហុស​មូលដ្ឋាន​ទិន្នន័យ: {err}'}), 500

    finally:
        cursor.close()
        connection.close()




# delet timetable
@timetable_bp.route('/<int:timetable_id>', methods=['DELETE'])
def delete_timetable(timetable_id):
    """Delete a timetable entry by ID."""
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # Check if the timetable entry exists
        cursor.execute("SELECT * FROM Timetable WHERE id = %s", (timetable_id,))
        timetable = cursor.fetchone()

        if not timetable:
            return jsonify({'error': 'Timetable entry not found'}), 404

        # Perform the delete operation
        cursor.execute("DELETE FROM Timetable WHERE id = %s", (timetable_id,))
        connection.commit()

        return jsonify({'message': 'Timetable entry deleted successfully'}), 200

    except mysql.connector.Error as err:
        connection.rollback()
        app.logger.error(f"Error deleting timetable: {err}")
        return jsonify({'error': f'An error occurred: {err}'}), 500

    finally:
        cursor.close()
        connection.close()
        
        
        
        
        
        
        


#delet timetable mysql
# @timetable_bp.route('/delete1', methods=['DELETE'])
# def delete_timetable1():
#     """Delete timetable entries by filters or by specific ID."""
#     connection = get_db_connection()
#     cursor = connection.cursor()

#     # Retrieve optional query parameters
#     year = request.args.get('year')
#     semester = request.args.get('semester')
#     major_id = request.args.get('major_id')
#     group_student = request.args.get('group_student')
#     generation = request.args.get('generation')

#     try:
#         # Build the DELETE query dynamically based on the filters provided      
#         query = "DELETE FROM Timetable WHERE 1=1"  # Starting point for dynamic query
#         params = []
        
#         if year:
#             query += " AND years = %s"
#             params.append(year)

#         if semester:
#             query += " AND semester = %s"
#             params.append(semester)
            
#         if generation:
#             query += " AND generation = %s"
#             params.append(generation)
            
#         if major_id:
#             query += " AND major_id = %s"
#             params.append(major_id)

#         if group_student:
#             query += " AND group_student = %s"
#             params.append(group_student)
            
            

#         # Check if any filter is provided, otherwise return an error
#         if not params:
#             return jsonify({'error': 'No filters provided for deletion'}), 400
        
#         # Execute the DELETE operation
#         cursor.execute(query, tuple(params))
#         connection.commit()

#         # Return a success message indicating the number of rows deleted
#         return jsonify({'message': f'Timetable entries deleted successfully'}), 200

#     except mysql.connector.Error as err:
#         connection.rollback()
#         app.logger.error(f"Error deleting timetable: {err}")
#         return jsonify({'error': f'An error occurred: {err}'}), 500

#     finally:
#         cursor.close()
#         connection.close()



# sql   server------
@timetable_bp.route('/delete1', methods=['DELETE'])
def delete_timetable1():
    """Delete timetable entries by filters or by specific ID."""
    connection = get_db_connection()
    cursor = connection.cursor()

    year = request.args.get('year')
    semester = request.args.get('semester')
    major_id = request.args.get('major_id')
    group_student = request.args.get('group_student')
    generation = request.args.get('generation')

    try:
        # Dynamic query
        query = "DELETE FROM Timetable WHERE 1=1"
        params = []
        
        if year:
            query += " AND years = ?"
            params.append(year)
            
        if semester:
            query += " AND semester = ?"
            params.append(semester)

        if generation:
            query += " AND generation = ?"
            params.append(generation)

        if major_id:
            query += " AND major_id = ?"
            params.append(major_id)

        if group_student:
            query += " AND group_student = ?"
            params.append(group_student)

        if not params:
            return jsonify({'error': 'No filters provided for deletion'}), 400

        cursor.execute(query, params)   # pyodbc accept list/tuple
        connection.commit()

        return jsonify({'message': f'Timetable entries deleted successfully'}), 200

    except Exception as err:
        connection.rollback()
        return jsonify({'error': f'An error occurred: {err}'}), 500

    finally:
        cursor.close()
        connection.close()

        




















# dont delet it funtion boot telegram  
#-1002688585037     7890827748:AAGtFzuC3l2YSsFUt-ijm0RInEYg-BN0lC4
#-1002577698834     mmo
#-1002632353520    me  
# now telegram send text 
def send_telegram_message(message):
    """Send a message to Telegram group"""
    TOKEN = "7890827748:AAGtFzuC3l2YSsFUt-ijm0RInEYg-BN0lC4"
    CHAT_ID = "-1002688585037"
    url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
    payload = {
        "chat_id": CHAT_ID , 
        "text": message, 
        "parse_mode": "Markdown"
        }
    response = requests.post(url, json=payload)
    
    if response.status_code == 200:
        send_telegram_message_optional("hi" ,chat_id='-1002565851640')
        return True
    else:
        return False
    
    
    
    
    
    
    

# 🔍 Check
@timetable_bp.route('/message_telegram', methods=['POST'])
def message_telegram():
    """Send a message to Telegram group"""
    data = request.json
    message = data.get('message')

    if not message:
        return jsonify({'error': 'Message is required'}), 400
 
    if send_telegram_message_optional(message):
        return jsonify({'message': 'Message sent successfully'}), 200
    else:
        return jsonify({'error': 'Error sending message'}), 500
    







# Route to trigger sending message to Telegram
@timetable_bp.route('tessx', methods=['GET', 'POST'])
def message_telegram1():
    """Send a message to Telegram group with a dynamic badge"""
    if request.method == 'POST':
        # POST request logic
        data = request.json
        badge_content = data.get('badgeContent')

        if not badge_content:
            return jsonify({'error': 'badgeContent is required'}), 400

        message = f"![Static Badge](https://img.shields.io/badge/{badge_content})"

        if send_telegram_message_optional(message):
            return jsonify({'message': 'Message sent successfully'}), 200
        else:
            return jsonify({'error': 'Error sending message'}), 500

    elif request.method == 'GET':
        # Send the Telegram message automatically when GET request is made
        message = """
 
```python\ndef x():\n    print\\(\\\"OK\\\"\\)\n```
AAAA ` enlight piece ` BBB
        
        """

        

        if send_telegram_message_optional(message):
            return jsonify({'message': 'Telegram message sent successfully via GET request'}), 200
        else:
            return jsonify({'error': 'Error sending message'}), 500



    
def send_telegram_message_optional(message, chat_id="-1002688585037", token="7890827748:AAGtFzuC3l2YSsFUt-ijm0RInEYg-BN0lC4"):
    """Send a message to a Telegram group with optional chat_id and token"""
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {
        "chat_id": chat_id,
        "text": message,
        "parse_mode": "Markdown" 
    }

    response = requests.post(url, json=payload)
    
    if response.status_code == 200:
        print("Message sent successfully!")
        return True
    else:
        print(f"Error: {response.text}")
        return False


def send_telegram_message_old(message):
    """Send a message to Telegram group"""
    url = f"https://api.telegram.org/bot7894181891:AAGqaMQ-uU1-GMFZYcP1EaDKvIVVGPVCmSQ/sendMessage"
    payload = {"chat_id": -1002632353520, "text": message, "parse_mode": "Markdown"}
    response = requests.post(url, json=payload)
    if response.status_code == 200:
        return True
    else:
        return False





    
    









