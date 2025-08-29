import os
from tempfile import TemporaryDirectory
from tkinter.font import Font
import traceback
import zipfile
from flask import Blueprint, Response, app, jsonify, logging, render_template, render_template_string, request
import mysql.connector
import openpyxl
import pdfkit
from config import get_db_connection
from flask import jsonify, request
from flask import current_app

timetable_bp = Blueprint('timetable', __name__, url_prefix='/api/timetable')


#  all this for make web to smoot not low new not used 
# for studen
#used http://127.0.0.1:5000/api/timetable/students?years=2025&semester=1&shift_name=Monday-Friday%20Morning&generation=16&group_student=1
# add &major_id=1

#http://127.0.0.1:5000/api/timetable/students?years=2025&semester=1&shift_name=Monday-Friday%20Morning&generation=16&group_student=1

@timetable_bp.route('/students', methods=['GET'])
def get_timetable_filtered_students():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Get query parameters
        years = request.args.get('years')
        semester = request.args.get('semester')
        shift_name = request.args.get('shift_name')
        generation = request.args.get('generation')
        group_student = request.args.get('group_student')
        major_id = request.args.get('major_id')

        # Ensure all required parameters are provided
        if not all([years, semester, shift_name, generation, group_student, major_id]):
            return jsonify({'error': 'សូមបញ្ជាក់គ្រប់ parameter ទាំងអស់: years, semester, shift_name, generation, group_student, major_id'}), 400

        # Base query
        query = """
            SELECT 
                t.id, 
                t.note, 
                t.study_sessions_id, 
                t.group_student, 
                t.batch, 
                t.generation, 
                t.major_id, 
                t.teacher_id, 
                t.subject_id, 
                t.room_id,
                t.years,
                t.semester,
                TIME_FORMAT(session_time_start, '%H:%i:%s') AS session_time_start,
                TIME_FORMAT(session_time_end, '%H:%i:%s') AS session_time_end,
                ss.shift_name AS study_shift_name,  
                ss.sessions_day AS study_session_day,
                m.name AS major_name,
                ts.teacher_id AS teacher_id, 
                ts.subject_id AS subject_id,
                r.room_number AS room_number,
                sub.name AS subject_name
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN teacher_subjects ts ON t.teacher_id = ts.teacher_id AND t.subject_id = ts.subject_id
            LEFT JOIN Rooms r ON t.room_id = r.id
            LEFT JOIN Subjects sub ON t.subject_id = sub.id
            WHERE 
                t.years = %(years)s AND 
                t.semester = %(semester)s AND 
                ss.shift_name = %(shift_name)s AND 
                t.generation = %(generation)s AND 
                t.group_student = %(group_student)s AND 
                t.major_id = %(major_id)s
        """

        # Parameters for the query
        params = {
            'years': years,
            'semester': semester,
            'shift_name': shift_name,
            'generation': generation,
            'group_student': group_student,
            'major_id': major_id
        }

        cursor.execute(query, params)
        timetable_entries = cursor.fetchall()

        return jsonify({'timetable_entries': timetable_entries})

    except Exception as e:
        logging.error(f"Error fetching timetable data: {e}")
        return jsonify({'error': 'មានបញ្ហាក្នុងការទាញយកទិន្នន័យ'}), 500
    finally:
        cursor.close()
        connection.close()
#http://127.0.0.1:5000/api/timetable/teacher?years=2025&semester=1&teacher_id=1
@timetable_bp.route('/teacher', methods=['GET'])
def get_timetable_filtered_teacher():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)
    try:
        # Get query parameters
        years = request.args.get('years')
        semester = request.args.get('semester')
        teacher_id = request.args.get('teacher_id')

        # Ensure all required parameters are provided
        if not all([years, semester, teacher_id]):
            return jsonify({'error': 'សូមបញ្ជាក់គ្រប់ parameter ទាំងអស់: years, semester, teacher_id'}), 400

        # Base query
        query = """
            SELECT 
                t.id, 
                t.note, 
                t.study_sessions_id, 
                t.group_student, 
                t.batch, 
                t.generation, 
                t.major_id, 
                t.teacher_id, 
                t.subject_id, 
                t.room_id,
                t.years,
                t.semester,
                TIME_FORMAT(session_time_start, '%H:%i:%s') AS session_time_start,
                TIME_FORMAT(session_time_end, '%H:%i:%s') AS session_time_end,
                ss.shift_name AS study_shift_name,  
                ss.sessions_day AS study_session_day,
                m.name AS major_name,
                ts.teacher_id AS teacher_id, 
                ts.subject_id AS subject_id,
                r.room_number AS room_number,
                sub.name AS subject_name
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN teacher_subjects ts ON t.teacher_id = ts.teacher_id AND t.subject_id = ts.subject_id
            LEFT JOIN Rooms r ON t.room_id = r.id
            LEFT JOIN Subjects sub ON t.subject_id = sub.id
            WHERE 
                t.years = %(years)s AND 
                t.semester = %(semester)s AND 
                t.teacher_id = %(teacher_id)s
        """

        params = {
            'years': years,
            'semester': semester,
            'teacher_id': teacher_id
        }

        cursor.execute(query, params)
        timetable_entries = cursor.fetchall()

        return jsonify({'timetable_entries': timetable_entries})

    except Exception as e:
        logging.error(f"Error fetching timetable data: {e}")
        return jsonify({'error': 'មានបញ្ហាក្នុងការទាញយកទិន្នន័យ'}), 500
    finally:
        cursor.close()
        connection.close()



@timetable_bp.route('/years', methods=['GET'])
def get_timetable_years():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch all unique years dynamically
        cursor.execute("""
            SELECT DISTINCT t.years 
            FROM Timetable t
            ORDER BY t.years ASC
        """)
        timetable_entries = cursor.fetchall()

        return jsonify({'years': timetable_entries})

    except Exception as e:
        app.logger.error(f"Error fetching timetable data: {e}")
        return jsonify({'error': 'An error occurred while fetching timetable data'}), 500
    finally:
        cursor.close()
        connection.close()


@timetable_bp.route('/semesters', methods=['GET'])
def get_timetable_semesters():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch all unique semesters dynamically
        cursor.execute("""
            SELECT DISTINCT t.semester 
            FROM Timetable t
            ORDER BY t.semester ASC
        """)
        timetable_entries = cursor.fetchall()

        return jsonify({'semesters': timetable_entries})

    except Exception as e:
        app.logger.error(f"Error fetching timetable semester data: {e}")
        return jsonify({'error': 'An error occurred while fetching timetable semester data'}), 500
    finally:
        cursor.close()
        connection.close()
        

@timetable_bp.route('/group_student', methods=['GET'])
def get_timetable_group_student():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch all unique semesters dynamically
        cursor.execute("""
            SELECT DISTINCT t.group_student
            FROM Timetable t
            ORDER BY t.group_student ASC
        """)
        timetable_entries = cursor.fetchall()

        return jsonify({'group_student': timetable_entries})

    except Exception as e:
        app.logger.error(f"Error fetching timetable semester data: {e}")
        return jsonify({'error': 'An error occurred while fetching timetable semester data'}), 500
    finally:
        cursor.close()
        connection.close()
        
@timetable_bp.route('/batch', methods=['GET'])
def get_timetable_batch():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch all unique semesters dynamically
        cursor.execute("""
            SELECT DISTINCT t.batch
            FROM Timetable t
            ORDER BY t.batch ASC
        """)
        timetable_entries = cursor.fetchall()

        return jsonify({'batch': timetable_entries})

    except Exception as e:
        app.logger.error(f"Error fetching timetable semester data: {e}")
        return jsonify({'error': 'An error occurred while fetching timetable semester data'}), 500
    finally:
        cursor.close()
        connection.close()
        
        
@timetable_bp.route('/generation', methods=['GET'])
def get_timetable_generation():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch all unique semesters dynamically
        cursor.execute("""
            SELECT DISTINCT t.generation
            FROM Timetable t
            ORDER BY t.generation ASC
        """)
        timetable_entries = cursor.fetchall()

        return jsonify({'generation': timetable_entries})

    except Exception as e:
        app.logger.error(f"Error fetching timetable semester data: {e}")
        return jsonify({'error': 'An error occurred while fetching timetable semester data'}), 500
    finally:
        cursor.close()
        connection.close()


        
@timetable_bp.route('/shift_name', methods=['GET'])
def get_shift_names():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch all unique shift_name values from study_sessions used in Timetable
        cursor.execute("""
            SELECT DISTINCT ss.shift_name
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            ORDER BY ss.shift_name ASC
        """)
        shift_names = cursor.fetchall()

        return jsonify({'shift_names': shift_names})

    except Exception as e:
        app.logger.error(f"Error fetching shift names: {e}")
        return jsonify({'error': 'An error occurred while fetching shift names'}), 500
    finally:
        cursor.close()
        connection.close()



@timetable_bp.route('/teacherid', methods=['GET'])
def get_teacher_details():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch teacher_id and teacher's name from the teachers table
        cursor.execute("""
            SELECT DISTINCT t.teacher_id, te.name AS teacher_name
            FROM Timetable t
            LEFT JOIN teachers te ON t.teacher_id = te.id
            ORDER BY t.teacher_id ASC
        """)
        teacher_details = cursor.fetchall()

        return jsonify({'teacher_details': teacher_details})

    except Exception as e:
        app.logger.error(f"Error fetching teacher details: {e}")
        return jsonify({'error': 'An error occurred while fetching teacher details'}), 500
    finally:
        cursor.close()
        connection.close()



from collections import defaultdict


@timetable_bp.route('auto', methods=['GET'])
def get_timetable_auto():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch timetable data with detailed error logging
        cursor.execute("""
            SELECT 
                t.id, 
                t.note, 
                t.study_sessions_id, 
                t.group_student, 
                t.batch, 
                t.generation, 
                t.major_id, 
                t.teacher_id, 
                t.subject_id, 
                t.room_id,
                t.years,
                t.semester,
                TIME_FORMAT(session_time_start, '%H:%i:%s') AS session_time_start,
                TIME_FORMAT(session_time_end, '%H:%i:%s') AS session_time_end,
                ss.shift_name AS study_shift_name,  
                ss.sessions_day AS study_session_day,
                m.name AS major_name,
                ts.teacher_id AS teacher_id, 
                ts.subject_id AS subject_id,
                r.room_number AS room_number,
                sub.name AS subject_name
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN teacher_subjects ts ON t.teacher_id = ts.teacher_id AND t.subject_id = ts.subject_id
            LEFT JOIN Rooms r ON t.room_id = r.id
            LEFT JOIN Subjects sub ON t.subject_id = sub.id
        """)
        timetable_entries = cursor.fetchall()

        # Group data by years > semester > major > generation > group_student
        grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list)))))

        for entry in timetable_entries:
            # Group by years > semester > major > generation > group_student
            grouped_data[entry['years']][entry['semester']][entry['major_id']][entry['generation']][entry['group_student']].append(entry)

        # Convert the grouped data to a JSON-compatible format
        # Nested dictionary iteration to make it JSON-compatible
        final_grouped_data = {}
        for year, year_data in grouped_data.items():
            final_grouped_data[year] = {}
            for semester, semester_data in year_data.items():
                final_grouped_data[year][semester] = {}
                for major_id, major_data in semester_data.items():
                    final_grouped_data[year][semester][major_id] = {}
                    for generation, generation_data in major_data.items():
                        final_grouped_data[year][semester][major_id][generation] = {}
                        for group_student, entries in generation_data.items():
                            final_grouped_data[year][semester][major_id][generation][group_student] = entries

        return jsonify({'grouped_timetable': final_grouped_data})

    except Exception as e:
        # Log the exception details
        app.logger.error(f"Error fetching timetable data: {e}")
        return jsonify({'error': 'An error occurred while fetching timetable data'}), 500
    finally:
        cursor.close()
        connection.close()




import pandas as pd
from io import BytesIO
from flask import send_file

@timetable_bp.route('/export_to_excel', methods=['GET'])
def export_to_excel():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch timetable data
        cursor.execute("""
            SELECT 
                t.id, 
                t.note, 
                t.study_sessions_id, 
                t.group_student, 
                t.batch, 
                t.generation, 
                t.major_id, 
                t.teacher_id, 
                t.subject_id, 
                t.room_id,
                t.years,
                t.semester,
                TIME_FORMAT(session_time_start, '%H:%i:%s') AS session_time_start,
                TIME_FORMAT(session_time_end, '%H:%i:%s') AS session_time_end,
                ss.shift_name AS study_shift_name,  
                ss.sessions_day AS study_session_day,
                m.name AS major_name,
                ts.teacher_id AS teacher_id, 
                ts.subject_id AS subject_id,
                r.room_number AS room_number,
                sub.name AS subject_name
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN teacher_subjects ts ON t.teacher_id = ts.teacher_id AND t.subject_id = ts.subject_id
            LEFT JOIN Rooms r ON t.room_id = r.id
            LEFT JOIN Subjects sub ON t.subject_id = sub.id
        """)
        timetable_entries = cursor.fetchall()

        # Convert timetable data to a flat structure suitable for Excel export
        rows = []
        for entry in timetable_entries:
            row = {
                "Year": entry['years'],
                "Semester": entry['semester'],
                "Generation": entry['generation'],
                "Major": entry['major_name'],
                "Group": entry['group_student'],
                "Batch": entry['batch'],
                "Subject": entry['subject_name'],
                "Teacher ID": entry['teacher_id'],
                "Teacher Name": entry.get('teacher_name', 'N/A'),  # Assuming teacher name is fetched or available
                "Room": entry['room_number'],
                "Study Shift": entry['study_shift_name'],
                "Session Day": entry['study_session_day'],
                "Session Time Start": entry['session_time_start'],
                "Session Time End": entry['session_time_end'],
                "Note": entry['note']
            }
            rows.append(row)

        # Convert rows to a DataFrame
        df = pd.DataFrame(rows)

        # Save DataFrame to Excel (in-memory)
        output = BytesIO()
        df.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)

        # Send file as response
        return send_file(output, as_attachment=True, download_name="timetable_data.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        app.logger.error(f"Error exporting timetable data to Excel: {e}")
        return jsonify({'error': 'An error occurred while exporting timetable data to Excel'}), 500
    finally:
        cursor.close()
        connection.close()



@timetable_bp.route('auto1', methods=['GET'])
def get_timetable_autoexcel1():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch timetable data with detailed error logging
        cursor.execute("""
            SELECT 
                t.id, 
                t.note, 
                t.study_sessions_id, 
                t.group_student, 
                t.batch, 
                t.generation, 
                t.major_id,
                m.name AS major_name, 
                t.teacher_id, 
                t.subject_id, 
                t.room_id,
                t.years,
                t.semester,
                TIME_FORMAT(session_time_start, '%H:%i:%s') AS session_time_start,
                TIME_FORMAT(session_time_end, '%H:%i:%s') AS session_time_end,
                ss.shift_name AS study_shift_name,  
                ss.sessions_day AS study_session_day,
                ts.teacher_id AS teacher_id, 
                ts.subject_id AS subject_id,
                r.room_number AS room_number,
                sub.name AS subject_name,
                te.name AS teacher_name  -- Fetch teacher name here
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN teacher_subjects ts ON t.teacher_id = ts.teacher_id AND t.subject_id = ts.subject_id
            LEFT JOIN Rooms r ON t.room_id = r.id
            LEFT JOIN Subjects sub ON t.subject_id = sub.id
            LEFT JOIN Teachers te ON t.teacher_id = te.id  -- Join Teachers table to get the teacher's name
        """)
        timetable_entries = cursor.fetchall()

        # Group data by years > semester > major > generation > group_student
        grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list)))))

        for entry in timetable_entries:
            # Group by years > semester > major > generation > group_student
            grouped_data[entry['years']][entry['semester']][entry['major_id']][entry['generation']][entry['group_student']].append(entry)
        rows = []
        


        # Store previous values to check for duplicates
        previous_values = {
            'year': None,
            'semester': None,
            'major': None,
            'generation': None,
            'group_student': None,
            'study_shift': None,  # Added study_shift to track its value
            'sessions_day': None  # Added sessions_day to track its value
        }

        # Loop through the grouped data
        for year, year_data in grouped_data.items():
            for semester, semester_data in year_data.items():
                for major_id, major_data in semester_data.items():
                    for generation, generation_data in major_data.items():
                        for group_student, entries in generation_data.items():

                            # Add empty row (optional)
                            rows.append({
                                'Year': '',
                                'Semester': '',
                                
                                'Major': '',
                                'Generation': '',

                                'Group Student': '',

                                'Study Shift': '',
                                'Sessions Day': '',

                     
                                'Subject Name': '',
                                'Teacher Name': '',
                                'Room Number': '',
                                
                                'Session Start': '',
                                'Session End': '',
                            })

                            # Loop through the timetable entries
                            for entry in entries:
                                # Check if the current row has the same value as the previous one for grouping
                                year_value = year if year != previous_values['year'] else ''
                                semester_value = semester if semester != previous_values['semester'] else ''
                                major_value = entry['major_name'] if entry['major_name'] != previous_values['major'] else ''
                                generation_value = generation if generation != previous_values['generation'] else ''
                                group_student_value = group_student if group_student != previous_values['group_student'] else ''
                                study_shift_value = entry['study_shift_name'] if entry['study_shift_name'] != previous_values['study_shift'] else ''
                                sessions_day_value = entry['study_session_day'] if entry['study_session_day'] != previous_values['sessions_day'] else ''
                    
                                # Append the row with merged columns if necessary
                                rows.append({
                                    'Year': year_value,
                                    'Semester': semester_value,
                                    
                                    'Major': major_value,
                                    'Generation': generation_value,
                                    'Group Student': group_student_value,
                    
                                    'Study Shift': study_shift_value,  # Using study_shift_value here
                                    'Sessions Day': sessions_day_value,  # Using sessions_day_value here
    
                                    'Subject Name': entry['subject_name'],
                                    'Teacher Name': entry['teacher_name'],  # Fetch teacher name
                                    'Room Number': entry['room_number'],
                                    
                                    'Session Start': entry['session_time_start'],
                                    'Session End': entry['session_time_end'],
                                })

                                # Update the previous values
                                previous_values['year'] = year
                                previous_values['semester'] = semester
                                previous_values['major'] = entry['major_name']
                                previous_values['generation'] = generation
                                previous_values['group_student'] = group_student
                                previous_values['study_shift'] = entry['study_shift_name']  # Update study_shift
                                previous_values['sessions_day'] = entry['study_session_day']  # Update sessions_day
        
                                        
                                
                                
        df = pd.DataFrame(rows)
        
        
        output = BytesIO()
        df.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="កាវិភាគទាំងអស់.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


      

    except Exception as e:
        # Log the exception details
        app.logger.error(f"Error fetching timetable data: {str(e)}")
        return jsonify({'error': f'An error occurred while fetching timetable data: {str(e)}'}), 500
    finally:
        cursor.close()
        connection.close()
        
      
      
      
@timetable_bp.route('/download', methods=['GET'])
def get_timetable_autoexcel2():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch timetable data with optimized query
        cursor.execute("""
            SELECT 
                t.id, 
                t.note, 
                t.study_sessions_id, 
                t.group_student, 
                t.batch, 
                t.generation, 
                t.major_id,
                m.name AS major_name, 
                t.teacher_id, 
                t.subject_id, 
                t.room_id,
                t.years,
                t.semester,
         TIME_FORMAT(session_time_start, '%H:%i:%s') AS session_time_start,
                TIME_FORMAT(session_time_end, '%H:%i:%s') AS session_time_end,
                ss.shift_name AS study_shift_name,  
                ss.sessions_day AS study_session_day,
                r.room_number AS room_number,
                sub.name AS subject_name,
                te.name AS teacher_name  -- Fetch teacher name directly
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN Rooms r ON t.room_id = r.id
            LEFT JOIN Subjects sub ON t.subject_id = sub.id
            LEFT JOIN Teachers te ON t.teacher_id = te.id  -- Direct join for teacher name
        """)
        timetable_entries = cursor.fetchall()

        # Group data by years > semester > major > generation > group_student
        grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list)))))

        for entry in timetable_entries:
            grouped_data[entry['years']][entry['semester']][entry['major_id']][entry['generation']][entry['group_student']].append(entry)

        rows = []

        # Store previous values to check for duplicates
        previous_values = {
            'year': None,
            'semester': None,
            'major': None,
            'generation': None,
            'group_student': None,
            'study_shift': None,
            'sessions_day': None
        }

        # Loop through grouped data
        for year, year_data in grouped_data.items():
            for semester, semester_data in year_data.items():
                for major_id, major_data in semester_data.items():
                    for generation, generation_data in major_data.items():
                        for group_student, entries in generation_data.items():
                            # Add an empty row (optional) before each new group
                            rows.append({
                                'Year': '', 'Semester': '', 'Major': '', 'Generation': '',
                                'Group Student': '', 'Study Shift': '', 'Sessions Day': '',
                                'Subject Name': '', 'Teacher Name': '', 'Room Number': '',
                                'Session Start': '', 'Session End': ''
                            })

                            # Loop through the timetable entries
                            for entry in entries:
                                # Avoid repeating values for merged columns
                                year_value = year if year != previous_values['year'] else ''
                                semester_value = semester if semester != previous_values['semester'] else ''
                                major_value = entry['major_name'] if entry['major_name'] != previous_values['major'] else ''
                                generation_value = generation if generation != previous_values['generation'] else ''
                                group_student_value = group_student if group_student != previous_values['group_student'] else ''
                                study_shift_value = entry['study_shift_name'] if entry['study_shift_name'] != previous_values['study_shift'] else ''
                                sessions_day_value = entry['study_session_day'] if entry['study_session_day'] != previous_values['sessions_day'] else ''

                                # Append row
                                rows.append({
                                    'Year': year_value,
                                    'Semester': semester_value,
                                    'Major': major_value,
                                    'Generation': generation_value,
                                    'Group Student': group_student_value,
                                    'Study Shift': study_shift_value,
                                    'Sessions Day': sessions_day_value,
                                    'Subject Name': entry['subject_name'],
                                    'Teacher Name': entry['teacher_name'],
                                    'Room Number': entry['room_number'],
                                    'Session Start': entry['session_time_start'],
                                    'Session End': entry['session_time_end'],
                                })

                                # Update previous values
                                previous_values.update({
                                    'year': year,
                                    'semester': semester,
                                    'major': entry['major_name'],
                                    'generation': generation,
                                    'group_student': group_student,
                                    'study_shift': entry['study_shift_name'],
                                    'sessions_day': entry['study_session_day']
                                })

        # Convert to DataFrame
        df = pd.DataFrame(rows)

        # Write to Excel file
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Timetable")
        
        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="កាវិភាគទាំងអស់.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        app.logger.error(f"Error fetching timetable data: {str(e)}")
        return jsonify({'error': f'An error occurred while fetching timetable data: {str(e)}'}), 500
    finally:
        cursor.close()
        connection.close()  
      
import pandas as pd
import zipfile
from io import BytesIO
from flask import send_file, jsonify

@timetable_bp.route('/downloade', methods=['GET'])
def get_timetable_autoexcel3():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch timetable data
        cursor.execute("""
            SELECT 
                t.id, 
                t.note, 
                t.study_sessions_id, 
                t.group_student, 
                t.batch, 
                t.generation, 
                t.major_id,
                m.name AS major_name, 
                t.teacher_id, 
                t.subject_id, 
                t.room_id,
                t.years,
                t.semester,
                TIME_FORMAT(session_time_start, '%H:%i:%s') AS session_time_start,
                TIME_FORMAT(session_time_end, '%H:%i:%s') AS session_time_end,
                ss.shift_name AS study_shift_name,  
                ss.sessions_day AS study_session_day,
                r.room_number AS room_number,
                sub.name AS subject_name,
                te.name AS teacher_name
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN Rooms r ON t.room_id = r.id
            LEFT JOIN Subjects sub ON t.subject_id = sub.id
            LEFT JOIN Teachers te ON t.teacher_id = te.id
        """)
        timetable_entries = cursor.fetchall()

        if not timetable_entries:
            return jsonify({'message': 'No timetable data found'}), 404

        # Group data by years > semester > major > generation > group_student
        grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list)  ))))
        
        for entry in timetable_entries:
            grouped_data[entry['years']][entry['semester']][entry['major_id']][entry['generation']][entry['group_student']].append(entry)

        # Prepare an in-memory ZIP file
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            
            # Loop through groups to create separate Excel files
            for year, year_data in grouped_data.items():
                for semester, semester_data in year_data.items():
                    for major_id, major_data in semester_data.items():
                        for generation, generation_data in major_data.items():
                            for group_student, entries in generation_data.items():
                                rows = []
                                previous_values = {
                                    'year': None,
                                    'semester': None,
                                    'major': None,
                                    'generation': None,
                                    'group_student': None,
                                    'study_shift': None,
                                    'sessions_day': None
                                }
                                for entry in entries:
                                    year_value = year if year != previous_values['year'] else ''
                                    semester_value = semester if semester != previous_values['semester'] else ''
                                    major_value = entry['major_name'] if entry['major_name'] != previous_values['major'] else ''
                                    generation_value = generation if generation != previous_values['generation'] else ''
                                    group_student_value = group_student if group_student != previous_values['group_student'] else ''
                                    study_shift_value = entry['study_shift_name'] if entry['study_shift_name'] != previous_values['study_shift'] else ''
                                    sessions_day_value = entry['study_session_day'] if entry['study_session_day'] != previous_values['sessions_day'] else ''

                                    
                                
                                    rows.append({
                                       
                     
                                        'Session Start': entry['session_time_start'],
                                        'Session End': entry['session_time_end'],
                                        'Sessions Day': sessions_day_value,
                                        'Subject Name': entry['subject_name'],
                                        'Teacher Name': entry['teacher_name'],
                                        'Room Number': entry['room_number'],
    
                                    })

                                    previous_values.update({
                                        'study_shift': entry['study_shift_name'],
                                        'sessions_day': entry['study_session_day'],
                                        'semester': semester,
                                        'major': entry['major_name'],
                                        'generation': generation,
                                        'group_student': group_student,
                                    
                                    })
                                    
                   
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                
                                

                                
                                                                # Initialize HTML content with CSS styling
                                html_content = """
                                <html>
                                <html lang="en">
                                <head>
                                    <meta charset="UTF-8">
                                    <meta name="viewport" content="width=device-width, initial-scale=1.0">
                             
                                </head>
                                
                                       <style>
                                        body {
                                                            font-family: Moul;
                                                            font-size: xx-small;
                                                        }
                                                        table {
                                                            width: 100%;
                                                            border-collapse: collapse;
                                                            font-family: Battambang;
                                                        }
                                                        th, td {
                                                            border: 1px solid black;
                                                            padding: 10px;
                                                            text-align: center;
                                                        }
                                                        .header {
                                                            text-align: center;
                                                            margin-bottom: 20px;
                                            }
                                    </style>
                                
                                
                                
                                
                                <div class="header" style=" display: flex;">


                                    <div  style="width: 20%;">
                                    
                                        <div>
                                            <br>
                                            <img src="https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcQEan2MdKi7N57D2ZeaZ7Y3iUauNq99Tdopcw&s"
                                                alt="University Logo" style="width:100px;">
                                        </div>
                                    
                                        <div>
                                            <p>សាកលវិទ្យាល័យជាតិមានជ័យ
                                                <br>
                                                ការិយាល័យសិក្សា
                                            </p>
                                        </div>
                                    </div>

                                
                                                                                
                                  <br>             
                                                    
                                                    
                            
                                """
                                
                                html_content +="""
                                          <div style="background-color: rgba(0, 255, 255, 0); width: 100%; transform: translateX(-10%);">
                                                ព្រះរាជាណាចក្រកម្ពុជា
                                                <br>
                                                ជាតិ សាសនា ព្រះមហាក្សត្រ

                                            
                                                <br>
                                                <img width="100px" src="https://puyvongheng.github.io/img/Screenshot%202024-11-13%20105603.png" alt="">
                                                <br>
                                                <br>
                                                <br>
                                                <br>
                                                
                   
                                                
                                                
                                                
                                     
                                """
                                
                                
                            #    html_content += f"<p>      ជំនាន់ទី {generation}, ក្រុមទី {group_student}    ឆ្នាំ  {year} កាលវិភាគប្រចាំឆមាសទី {semester}, វេនសិក្សា ជំនាញ  {major_value}, "

                              
                                
                                html_content += f"""<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">​​   ​</span>
                                """
                                html_content += f"  ឆ្នាំសិក្សា   {year} - {year +1} កាលវិភាគប្រចាំឆមាសទី {semester},  "
                                
                                html_content += f"""<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">​​ <br>  ​</span>
                                """
                                
                                html_content += f"""<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">​​ ចាប់ផ្តើមពីថ្ងៃ....កើត ខែ.... ឆ្នាំ.... .... ត្រូវនឹងថ្ងៃទី.... ខែ....ឆ្នាំ.... វេនសិក្សា .... ​</span>
                                """
                                
                                html_content += f"""<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">​​ <br>  ​</span>
                                """

                                html_content += f"  ជំនាញ{entry['major_name']}   ជ {generation}   {entry['study_shift_name']}  </p> "


                                html_content += """
                                
        
                                               
                
            
                                                
                                            </div>
                                            
                                    </div>
                                    <!--  vvvv-->
                                    <br>
                                """
                                
                                
                                
                                # Check if study_shift_name is "Saturday-Sunday" or not
                                html_content += """
                                    <body style="margin: 10px;">
                                        <table>
                                            <thead id="timetable">
                                                <tr>
                                                    <th>ម៉ោងសិក្សា</th>
                                """

                                # Check if study_shift_name is "Saturday-Sunday" or not
                                if any(entry['study_shift_name'].lower() != 'saturday-sunday' for entry in entries):
                                    # Show Monday to Friday columns
                                    html_content += """
                                                    <th>ថ្ងៃច័ន្ទ/Monday</th>
                                                    <th>ថ្ងៃអង្គារ/Tuesday</th>
                                                    <th>ថ្ងៃពុធ/Wednesday</th>
                                                    <th>ថ្ងៃព្រហស្បតិ៍/Thursday</th>
                                                    <th>ថ្ងៃសុក្រ/Friday</th>
                                                """
                                else:
                                    # Show Saturday and Sunday columns
                                    html_content += """
                                                    <th>ថ្ងៃសៅរ៍/Saturday</th>
                                                    <th>ថ្ងៃអាទិត្យ/Sunday</th>
                                                """
                                html_content += """
                                            </tr>
                                        </thead>
                                        <tbody>
                                """

                                # Dictionary to group sessions by time
                                grouped_sessions = {}

                                # Group sessions by session_time_start and session_time_end
                                for entry in entries:
                                    session_time = f"{entry['session_time_start']} - {entry['session_time_end']}"

                                    if session_time not in grouped_sessions:
                                        grouped_sessions[session_time] = {
                                            'monday': '',
                                            'tuesday': '',
                                            'wednesday': '',
                                            'thursday': '',
                                            'friday': '',
                                            'saturday': '',
                                            'sunday': ''
                                        }

                                    # Add subject details for the corresponding day
                                    day = entry['study_session_day'].lower()  # Convert day to lowercase for keys (monday, tuesday, etc.)

                                    subject_details = f"{entry['subject_name']}<br>{entry['study_shift_name']}<br>{entry['teacher_name']}<br>{entry['room_number']}"

                                    # Append subject details to the correct day slot
                                    if day == 'monday':
                                        grouped_sessions[session_time]['monday'] += subject_details + "<br>"
                                    elif day == 'tuesday':
                                        grouped_sessions[session_time]['tuesday'] += subject_details + "<br>"
                                    elif day == 'wednesday':
                                        grouped_sessions[session_time]['wednesday'] += subject_details + "<br>"
                                    elif day == 'thursday':
                                        grouped_sessions[session_time]['thursday'] += subject_details + "<br>"
                                    elif day == 'friday':
                                        grouped_sessions[session_time]['friday'] += subject_details + "<br>"
                                    elif day == 'saturday':
                                        grouped_sessions[session_time]['saturday'] += subject_details + "<br>"
                                    elif day == 'sunday':
                                        grouped_sessions[session_time]['sunday'] += subject_details + "<br>"

                                # Generate table rows from grouped sessions
                                row_count = 0

                                # Flag to check if we need to limit rows
                                limit_rows = any(entry['study_shift_name'].lower() != 'saturday-sunday' for entry in entries)

                                # Generate table rows from grouped sessions
                                for session_time, days in grouped_sessions.items():
                                    # If limit_rows is True and we have already added 3 rows, break the loop
                                    if limit_rows and row_count >= 3:
                                        break

                                    # Start the row with session time
                                    html_content += f"<tr><td>{session_time}</td>"

                                    # Add Monday to Friday columns if condition is met
                                    if any(entry['study_shift_name'].lower() != 'saturday-sunday' for entry in entries):
                                        html_content += f"<td>{days['monday']}</td>"
                                        html_content += f"<td>{days['tuesday']}</td>"
                                        html_content += f"<td>{days['wednesday']}</td>"
                                        html_content += f"<td>{days['thursday']}</td>"
                                        html_content += f"<td>{days['friday']}</td>"
                                    else:
                                        # Add Saturday and Sunday columns if "Saturday-Sunday" shift is found
                                        html_content += f"<td>{days['saturday']}</td>"
                                        html_content += f"<td>{days['sunday']}</td>"

                                    # End the row
                                    html_content += "</tr>"

                                    # Increment the row counter
                                    row_count += 1

                                # Close the table and body
                                html_content += """
                                    </tbody>
                                </table>
                                </body>


                                
                                
                                
                                
                  
                                
                                            <div style="display: flex; font-family: Battambang;">
                                                    <div class="footer" style="text-align: center; margin-top: 20px; width: 50%; ">
                                                        <p style=" font-family: Moul;">
                                                        <br>
                                                        <br>
                                                        <br><br><br>
                                                        <span style="font-family: Battambang;">
                                                        បានឃើញ និងឯកភាព
                                                        </span>
                                                        <br>
                                                        ជ.សាកលវិទ្យាធិការ
                                                        <br>
                                                        សាកលវិទ្យាធិការរង
                                                        <br>
                                                        </p>
                                                    </div>

                                                <div class="footer" style="text-align: center; margin-top: 20px; width: 50%;">
                                                    <p>ថ្ងៃទី<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">.......</span> ខែ<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">.......</span> ឆ្ន<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">.......</span> <span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">.......</span></p>
                                                    <p >បន្ទាយមានជ័យ ថ្ងៃទី<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">.......</span> ខែ<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">.......</span> ឆ្នាំ<span contenteditable="true" style="border-bottom: 1px  rgb(0, 0, 0);">.......</span></p>
                                                    <p>ប្រធានការិយាល័យសិក្សា</p>
                                                </div>



                                            </div>

                                </html>
                                """

                                # Add HTML file to ZIP
                                filename_html = f"{entry['study_shift_name']} ឆ្នាំ{year}__ ឆមាស{semester}__{entry['major_name']} ជំាញ {major_id}{major_value}__ ជំានាន{generation}__ ក្រុម{group_student}.html"
                                zip_file.writestr(filename_html, html_content)

                                                                
    
                    
                    
                                                    # Insert custom rows (header and filename)
                                # Add Big Title, Subtitle, and Header to the rows list
    # Insert Big Title and Subtitle into rows
                                rows.insert(0, {
                                    
                                    'year': 'ឆ្នាំ {} ឆមាស {}   ជំាញ {} ជំានាន {} ក្រុម {} {}'.format(year, semester , major_value, generation, group_student ,study_shift_value),
                                    'Session End': '',
                                    'Sessions Day': '',
                                    'Subject Name': '', 
                                    'Teacher Name': '', 
                                    'Room Number': '', 
                            
                                })

                                rows.insert(1, {
                                    'Session Start': '',
                                    'Session End': '',
                                    'Sessions Day':  '',
                                    'Subject Name': '', 
                                    'Teacher Name': '', 
                                    'Room Number': '', 
                                })

                                # Insert Column Headers
                                rows.insert(2, {
                            
                                    'Session Start': 'Session Start', 
                                    'Session End': 'Session End',
                                    'Sessions Day': 'Sessions Day', 
                                    'Subject Name': 'Subject Name', 
                                    'Teacher Name': 'Teacher Name', 
                                    'Room Number': 'Room Number', 
                                
                                })
                                # Convert to DataFrame
                                
                                
                                filename = f"ឆ្នាំ{year}__ ឆមាស{semester}__ ជំាញ{major_value}__ ជំានាន{generation}__ ក្រុម{group_student}.xlsx"
                              
                                df = pd.DataFrame(rows)
                                
                                
                           
                          
                                
                            
        
                                # Create in-memory Excel file
                                excel_buffer = BytesIO()
                                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                                    df.to_excel(writer, index=False, sheet_name="Timetable")
                                    
                                    
                                      # Get the workbook and worksheet
                                    worksheet = writer.sheets["Timetable"]
                                    
                         
                                    # Adjust column width based on the length of data in each column
                                    for col_num in range(2, worksheet.max_column + 1):  # Start from column 2 (B)
                                        column = worksheet.cell(row=1, column=col_num).column_letter  # Get the column letter (e.g. 'B', 'C', ...)
                                        max_length = 0
                                        for cell in worksheet[column]:
                                            try:
                                                if len(str(cell.value)) > max_length:
                                                    max_length = len(cell.value)
                                            except:
                                                pass
                                        adjusted_width = max_length + 2  # Add a little padding
                                        worksheet.column_dimensions[column].width = adjusted_width




                                    # Apply styles (e.g., bold headers)
                                    for cell in worksheet["1:1"]:
                                        cell.font = openpyxl.styles.Font(bold=True)
                                        
                                    worksheet.delete_rows(1)
                                    
                                
                                        
                                        
                                    
                                        # Apply style to row 2 (Subtitle Row) specifically
                                    for col_num in range(1, worksheet.max_column + 1):
                                        cell = worksheet.cell(row=3, column=col_num)
                                        cell.alignment = Alignment(horizontal='center', vertical='center')
                                        cell.font = Font(bold=True)


                                          
                                    fill_light_blue = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                                    for col_num in range(1, worksheet.max_column + 1):
                                        cell = worksheet.cell(row=3, column=col_num)
                                        cell.fill = fill_light_blue 
                                         
                                    for col_num in range(1, worksheet.max_column + 1):
                                        cell = worksheet.cell(row=1, column=col_num)
                                        cell.alignment = Alignment(horizontal='center', vertical='center')                                                                                                                           
                                        worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=worksheet.max_column)
                                        
                                        
                                        
                                    
                                                                      
   

                                excel_buffer.seek(0)

                                # Generate unique filename
                           
                                # Add to ZIP file
                                zip_file.writestr(filename, excel_buffer.getvalue())

        zip_buffer.seek(0)
        
        
        
        
        
   
        

        # Return ZIP file as response
        return send_file(
            zip_buffer,
          # as_attachment=True,
            download_name="Timetable_Files.zip",
            mimetype="application/zip",
        )

    except Exception as e:
        app.logger.error(f"Error fetching timetable data: {str(e)}")
        return jsonify({'error': f'An error occurred while fetching timetable data: {str(e)}'}), 500

    finally:
        cursor.close()
        connection.close()
        
from openpyxl.styles import PatternFill
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

















@timetable_bp.route('', methods=['GET'])
def get_timetable():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Fetch timetable data with detailed error logging
        cursor.execute("""
            SELECT 
                t.id, 
                t.note, 
                t.study_sessions_id, 
                t.group_student, 
                t.batch, 
                t.generation, 
                t.major_id, 
                t.teacher_id, 
                t.subject_id, 
                t.room_id,
                        t.years,
                        t.semester,
                       
                TIME_FORMAT(session_time_start, '%H:%i:%s') AS session_time_start,
                TIME_FORMAT(session_time_end, '%H:%i:%s') AS session_time_end,
                ss.shift_name AS study_shift_name,  
                ss.sessions_day AS study_session_day,
                m.name AS major_name,
                ts.teacher_id AS teacher_id, 
                ts.subject_id AS subject_id,
                r.room_number AS room_number,
                sub.name AS subject_name
                 
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN teacher_subjects ts ON t.teacher_id = ts.teacher_id AND t.subject_id = ts.subject_id
            LEFT JOIN Rooms r ON t.room_id = r.id
            LEFT JOIN Subjects sub ON t.subject_id = sub.id
         
        """)
        timetable_entries = cursor.fetchall()

        return jsonify({'timetable_entries': timetable_entries})

    except Exception as e:
        # Log the exception details
        app.logger.error(f"Error fetching timetable data: {e}")
        return jsonify({'error': 'An error occurred while fetching timetable data'}), 500
    finally:
        cursor.close()
        connection.close()
      

# 🔍 Check
@timetable_bp.route('/checkconflicts', methods=['POST'])
def check_timetable_conflicts():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    data = request.json
    required_fields = ['study_sessions_id', 'years', 'semester']
    missing_fields = [field for field in required_fields if not data.get(field)]
    if missing_fields:
        return jsonify({'error': f'Missing required fields: {missing_fields}'}), 400

    # Extract required parameters
    study_sessions_id = data.get('study_sessions_id')
    years = data.get('years')
    semester = data.get('semester')
    teacher_id = data.get('teacher_id')
    room_id = data.get('room_id')
    subject_id = data.get('subject_id')
    major_id = data.get('major_id')
    generation = data.get('generation')
    group_student = data.get('group_student')

    try:
        conflicts = []



        # បន្ថែមការពិនិត្យថ្មី៖ គ្រូម្នាក់មិនអាចបង្រៀនជំនាន់ខុសគ្នា នៅក្នុង study_sessions_id, years, និង semester
        cursor.execute("""
            SELECT id FROM Timetable 
            WHERE teacher_id = %s AND study_sessions_id = %s 
            AND years = %s AND semester = %s AND generation != %s
        """, (teacher_id, study_sessions_id, years, semester, generation))

        if cursor.fetchone():
            conflicts.append({'room': '🚫 គ្រូម្នាក់មិនអាចបង្រៀនជំនាន់ផ្សេងគ្នាក្នងម៉ោងសិក្សាដូចគ្នា!'})




        # Check for room conflicts
        cursor.execute("""
            SELECT id, room_id, study_sessions_id, years, semester 
            FROM Timetable 
            WHERE room_id = %s AND study_sessions_id = %s AND years = %s AND semester = %s
        """, (room_id, study_sessions_id, years, semester))
        room_conflict = cursor.fetchone()
        if room_conflict:
            conflicts.append({
                'room': '🚫 បន្ទប់នេះមិនទំនេរ!',
            })
        cursor.fetchall()  # Ensure we fetch all rows here to avoid unread result

        # Check for teacher conflicts
        cursor.execute("""
            SELECT id, teacher_id, study_sessions_id, years, semester 
            FROM Timetable 
            WHERE teacher_id = %s AND study_sessions_id = %s AND years = %s AND semester = %s
        """, (teacher_id, study_sessions_id, years, semester))
        teacher_conflict = cursor.fetchone()
        if teacher_conflict:
            conflicts.append({
                'teacher': '🚫 គ្រូរមិនទំនេរ!',
            })
        cursor.fetchall()  # Ensure we fetch all rows here to avoid unread result



        # Check for room and subject conflicts
        cursor.execute("""
            SELECT id FROM Timetable 
            WHERE room_id = %s AND teacher_id = %s AND years = %s AND semester = %s AND subject_id = %s AND study_sessions_id  = %s
        """, (room_id, teacher_id, years, semester, subject_id,study_sessions_id))
        room_subject_exists = cursor.fetchone()
        if room_subject_exists:
            conflicts.append({
                'room_subject': 'គ្រូបានដា់អោយវង្រៀនហើយ 🤝 អាចបន្ថែមសិស្សក្រុមផ្សេងបាន (+them)! ទោបីជា  បន្ទប់នេះមិនទំនេរ!',
            })
         #  else:
         #      conflicts.append({
         #          'room_subject': ' '
        #       })
            cursor.fetchall()  # Ensure we fetch all rows here to avoid unread result





 


       # Check if the teacher is already assigned to this room and subject in the same session
        cursor.execute("""
            SELECT 
                t.id, 
                t.group_student, 
                t.generation, 
                m.name AS major_name,
                r.room_number AS room
                
                
                
        
            FROM Timetable t
            JOIN majors m ON t.major_id = m.id
            JOIN rooms r ON t.room_id = r.id
 
            WHERE  t.teacher_id = %s 
            AND t.years = %s 
            AND t.semester = %s 
          
            AND t.study_sessions_id = %s
        """, ( teacher_id, years, semester,  study_sessions_id))

        room_subject_exists = cursor.fetchall()

        if room_subject_exists:
            conflict_messages = []
            for row in room_subject_exists:
                timetable_generation = row['generation']
                major_name = row['major_name']
                group = row['group_student']
                room = row['room']
                
                conflict_messages.append(f'- ជំនាន់ {timetable_generation} ជំនាញ ({major_name} {group}) បន្ទប់: {room}')
            
            # Join all conflict messages into a single line, separated by commas
            conflict_line = ', '.join(conflict_messages)
            
            # Append the formatted message to conflicts
            conflicts.append({
                'room_subject': f' គ្រូបានដាក់អោយបង្រៀនហើយ 🤝 អាចបន្ថែមសិស្សក្រុមផ្សេងបាន  {conflict_line}'
            })

        cursor.fetchall()  # Ensure we fetch all rows here to avoid unread result
                    
            





        # Check for group student conflicts
        cursor.execute("""
            SELECT * FROM Timetable 
            WHERE 
            group_student = %s 
            AND study_sessions_id = %s 
            AND years = %s 
            AND semester = %s 
            AND major_id = %s 
            AND generation = %s
        """, (group_student, study_sessions_id, years, semester, major_id, generation))
        group_student_conflict = cursor.fetchone()
        if group_student_conflict:
            conflicts.append({
                'messagesessions': '🚫 មានម៉ោងសិក្សារហើយ!'
            })
        cursor.fetchall()  # Ensure we fetch all rows here to avoid unread result



        # Check if the teacher is available at the study session
        cursor.execute("""
            SELECT * FROM teacher_teaching_time
            WHERE teacher_id = %s AND study_sessions_id = %s
        """, (teacher_id, study_sessions_id))
        if not cursor.fetchone():
            conflicts.append({'teacher_time': '🚫 គ្រូមិនអាចបង្រៀននៅម៉ោងនេះបានទេ'})
        cursor.fetchall()  # Ensure we fetch all rows here to avoid unread result
        

        # Return conflicts if found
        if conflicts:
            return jsonify({'conflicts': conflicts}), 200
        else:
            return jsonify({'message': 'No conflicts found'}), 200

    except mysql.connector.Error as err:
        # MySQL specific error logging
        print(f"MySQL Error: {err}")
        return jsonify({'error': f'MySQL Error: {err}'}), 500
    except Exception as e:
        # General error logging
        print(f"An error occurred: {str(e)}")
        return jsonify({'error': f'An error occurred: {str(e)}'}), 500
    finally:
        # Ensure all results are fetched before closing the cursor
        try:
            cursor.fetchall()  # Fetch any remaining results to avoid 'Unread result' error
        except Exception as e:
            print(f"Error while fetching remaining results: {str(e)}")
        cursor.close()  # Close the cursor
        connection.close()  # Close the connection


# create timetable 
@timetable_bp.route('', methods=['POST'])
def create_timetable():
    """✏️ បង្កើតកាលវិភាគថ្មី"""
    data = request.json
    current_app.logger.info(f"Received data: {data}")
    # 🛑 ពិនិត្យពាក្យចាំបាច់ 🛑
    required_fields = ['study_sessions_id', 'group_student', 'batch', 'generation',
                       'major_id', 'teacher_id', 'subject_id', 'room_id', 'years', 'semester']
    
    
    # Define a dictionary to map English field names to Khmer labels
    khmer_labels = {
        'study_sessions_id': 'វគ្គសិក្សា',
        'group_student': 'ក្រុមសិស្ស',
        'batch': 'ឆ្នាំ',
        'generation': 'ជំនាន់',
        'major_id': 'ជំនាញ',
        'teacher_id': 'គ្រូបង្រៀន',
        'subject_id': 'មុខវិជ្ជា',
        'room_id': 'បន្ទប់',
        'years': 'ឆ្នាំសិក្សា',
        'semester': 'ឆមាស'
    }

    # Find missing fields
    missing_fields = [khmer_labels[field] for field in required_fields if not data.get(field)]


   # missing_fields = [field for field in required_fields if not data.get(field)]
    errors = []
    if missing_fields:
        errors.append({'error': f"❌ ភ្លេច បញ្ចូល  <br>  <br>{' <br>  '.join(missing_fields)}"}), 400

    # 📌 ទាញយកព័ត៌មានពី Request
    note = data.get('note', '').strip()
    study_sessions_id = data.get('study_sessions_id')
    group_student = data.get('group_student')
    batch = data.get('batch')
    generation = data.get('generation')
    major_id = data.get('major_id')
    teacher_id = data.get('teacher_id')
    subject_id = data.get('subject_id')
    room_id = data.get('room_id')
    years = data.get('years')
    semester = data.get('semester')

    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        
        

        
        
    # បន្ថែមការពិនិត្យថ្មី៖ គ្រូម្នាក់មិនអាចបង្រៀនជំនាន់ខុសគ្នា នៅក្នុង study_sessions_id, years, និង semester
        cursor.execute("""
            SELECT id FROM Timetable 
            WHERE teacher_id = %s AND study_sessions_id = %s 
            AND years = %s AND semester = %s AND generation != %s
        """, (teacher_id, study_sessions_id, years, semester, generation))

        if cursor.fetchone():
            errors.append({'error': '🚫 គ្រូម្នាក់មិនអាចបង្រៀនជំនាន់ផ្សេងគ្នាក្នងម៉ោងសិក្សាដូចគ្នា!'})


        
        
        cursor.execute("SELECT number_sessions FROM Teachers WHERE id = %s", (teacher_id,))
        teacher = cursor.fetchone()
        if not teacher:
            errors.append({'error': '🚫 Teacher not found.'}), 400
            return jsonify({'errors': errors}), 400

        # Access number_sessions using index (0 for the first column)
        number_sessions = teacher[0]  # teacher is a tuple, so we access the first element

        # 🔍 Query to count the current timetable entries for the teacher
        cursor.execute("""
            SELECT COUNT(DISTINCT CONCAT(room_id, '-', study_sessions_id)) AS total_timetable
            FROM Timetable
            WHERE teacher_id = %s AND years = %s AND semester = %s
        """, (teacher_id, years, semester))

        result = cursor.fetchone()

        # Access total_timetable using index (0 for the first column in the tuple)
        total_timetable = result[0] if result else 0  # result is a tuple, so access the first element

        # 🔍 Check if total timetable exceeds teacher's allowed sessions
        if total_timetable >= number_sessions:
            errors.append({
                'error': f"🚫 គ្រូបានឈានដល់ចំនួនអតិបរមានៃ sessions ({number_sessions})ហើយសម្រាប់ឆមាសនេះ"
            })
            return jsonify({'errors': errors}), 400

                
        
        
  
        
        
        
        # 🔍 ពិនិត្យមើលថាបន្ទប់ត្រូវបានប្រើរួចឬអត់  ROOM
        cursor.execute("""
            SELECT t.id, ss.id AS study_sessions_id, t.years, t.semester, 
                m.name AS major_name, sub.name AS subject_name, te.name AS teacher_name
            FROM Timetable t
            LEFT JOIN study_sessions ss ON t.study_sessions_id = ss.id
            LEFT JOIN Majors m ON t.major_id = m.id
            LEFT JOIN Subjects sub ON t.subject_id = sub.id
            LEFT JOIN Teachers te ON t.teacher_id = te.id
            WHERE t.room_id = %s AND t.study_sessions_id = %s AND t.years = %s AND t.semester = %s
            AND NOT (t.teacher_id = %s AND t.subject_id = %s)
        """, (room_id, study_sessions_id, years, semester, teacher_id, subject_id))
        existing_room = cursor.fetchone()
        if existing_room:
            errors.append({
                'error': f'🚫 <strong>បន្ទប់មិនទំនេរ</strong> <br/>'
                        f'<strong>Timetable ID:</strong> {existing_room[0]} '
                        f'<strong>វគ្គសិក្សា:</strong> {existing_room[1]} '
                        f'<strong>ឆ្នាំសិក្សា:</strong> {existing_room[2]} '
                        f'<strong>ឆមាស:</strong> {existing_room[3]} <br/>'
                        f'<strong>ជំនាញ:</strong> {existing_room[4]} <br/>'
                        f'<strong>មុខវិជ្ជា:</strong> {existing_room[5]} <br/>'
                        f'<strong>គ្រូបង្រៀន:</strong> {existing_room[6]} <br/>​ <hr>'
            })






        # 🔍 ពិនិត្យមើលថាគ្រូស្ថិតនៅក្នុងបន្ទប់តែមួយក្នុង study_sessions_id, years, semester
        cursor.execute("""
            SELECT t.id, t.room_id, t.study_sessions_id, t.years, t.semester
            FROM Timetable t
            WHERE t.teacher_id = %s 
                AND t.study_sessions_id = %s 
                AND t.years = %s 
                AND t.semester = %s 
                AND t.room_id <> %s
        """, (teacher_id, study_sessions_id, years, semester, room_id))

        existing_teacher_room = cursor.fetchone()

        if existing_teacher_room:
            errors.append({
                'error': f'🚫 <strong>គ្រូបជាប់ង្រៀនហើយ ប្រហែលជានៅបន្ទប់ផ្សេង</strong> <br/>'
                        f'👨‍🏫 <strong>សម្រាប់sessionsនេះ</strong> <br/>'
                  
            })
            
            
            
            #មិនទានហើយ
        # 🔍 ពិនិត្យមើលថាមាន batch ដដែលអត់សម្រាប់ study_sessions_id, years, semester
  # 🔍 ពិនិត្យមើលថាតើមានកំណត់ត្រា Timetable សម្រាប់ study_sessions_id, years, semester និង generation ទេ
 

            
            
            
            
            
            
 






        # 🔍 ពិនិត្យថាក្រុមសិស្សមានម៉ោងសិក្សារួចនៅឬអត់  stusen  AND generation = %s
        cursor.execute("""
                SELECT * FROM Timetable 
                WHERE 
                group_student = %s AND 
                study_sessions_id = %s AND 
                years = %s AND 
                Semester = %s  AND 
                major_id = %s AND
                generation  = %s AND  
                batch = %s
            """, (group_student, study_sessions_id, years, semester, major_id,generation,batch))

        if cursor.fetchone():
            errors.append({
                    'error': '🚫 <strong>ក្រុមសិស្សរួចហើយ</strong> <br/>'
                            '👨‍🎓 <strong>សម្រាប់វគ្គនេះ</strong> <br/> <hr>'
            })
            
            
        
        
        
    

      
        







        # 🔍 ពិនិត្យមើលថាគ្រូនេះអាចបង្រៀនមុខវិជ្ជានេះឬអត់ teacher
        cursor.execute("""
            SELECT * FROM teacher_subjects 
            WHERE teacher_id = %s AND subject_id = %s
        """, (teacher_id, subject_id))

        if not cursor.fetchone():
            errors.append({'error': '🚫 គ្រូមិនអាចបង្រៀនមុខវិជ្ជានេះបានទេ ❌'}), 400

        # 🔍 ពិនិត្យមើលថាគ្រូមានសមត្ថភាពបង្រៀននៅម៉ោងនេះឬអត់ teacher
        cursor.execute("""
            SELECT * FROM teacher_teaching_time
            WHERE teacher_id = %s AND study_sessions_id = %s
        """, (teacher_id, study_sessions_id))

        if not cursor.fetchone():
            errors.append({'error': '🚫 គ្រូមិនអាចបង្រៀននៅម៉ោងនេះបានទេ ❌'}), 400
            
            
       
    
       
   
        
        

        # 👉 ប្រសិនបើមានកំហុសណាមួយ
        if errors:
            return jsonify({'errors': errors}), 400

        # ✅ បញ្ចូលទិន្នន័យកាលវិភាគថ្មី
        cursor.execute("""
            INSERT INTO Timetable (
                study_sessions_id, group_student, note, batch, generation, major_id,
                teacher_id, subject_id, room_id, years, Semester
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (study_sessions_id, group_student, note, batch, generation, major_id,
              teacher_id, subject_id, room_id, years, semester))

        connection.commit()
        return jsonify({'message': '🎉 បានបង្កើតដោយជោគជ័យ! ✅'}), 201

    except mysql.connector.Error as err:
        connection.rollback()
        current_app.logger.error(f"🚨 Error creating timetable: {err}")
        errors.append({'error': f'❌ មានបញ្ហា៖ {err}'}), 500
        return jsonify({'errors': errors}), 500
    finally:
        cursor.close()
        connection.close()


# delet timetable
@timetable_bp.route('/<int:timetable_id>', methods=['DELETE'])
def delete_timetable(timetable_id):
    """Delete a timetable entry by ID."""
    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # Check if the timetable entry exists
        cursor.execute("SELECT * FROM Timetable WHERE id = %s", (timetable_id,))
        timetable = cursor.fetchone()

        if not timetable:
            return jsonify({'error': 'Timetable entry not found'}), 404

        # Perform the delete operation
        cursor.execute("DELETE FROM Timetable WHERE id = %s", (timetable_id,))
        connection.commit()

        return jsonify({'message': 'Timetable entry deleted successfully'}), 200

    except mysql.connector.Error as err:
        connection.rollback()
        app.logger.error(f"Error deleting timetable: {err}")
        return jsonify({'error': f'An error occurred: {err}'}), 500

    finally:
        cursor.close()
        connection.close()
        













# Update a timetable 
@timetable_bp.route('/<int:timetable_id>', methods=['PUT'])
def update_timetable(timetable_id):
    """Update a timetable entry by ID."""
    data = request.json
    required_fields = [
        'study_sessions_id', 'group_student', 'batch', 'generation',
        'major_id', 'teacher_id', 'subject_id', 'room_id', 'years', 'semester'
    ]
    missing_fields = [field for field in required_fields if field not in data]
    if missing_fields:
        return jsonify({'error': f"Missing fields: {', '.join(missing_fields)}"}), 400

    connection = get_db_connection()
    cursor = connection.cursor()

    try:
        # Check if timetable entry exists
        cursor.execute("SELECT * FROM Timetable WHERE id = %s", (timetable_id,))
        existing_entry = cursor.fetchone()
        if not existing_entry:
            return jsonify({'error': 'Timetable entry not found'}), 404

        # Extract fields for update
        study_sessions_id = data.get('study_sessions_id')
        group_student = data.get('group_student')
        batch = data.get('batch')
        generation = data.get('generation')
        major_id = data.get('major_id')
        teacher_id = data.get('teacher_id')
        subject_id = data.get('subject_id')
        room_id = data.get('room_id')
        years = data.get('years')
        semester = data.get('semester')
        note = data.get('note', '').strip()

        # Update the timetable entry
        cursor.execute("""
            UPDATE Timetable
            SET 
                study_sessions_id = %s,
                group_student = %s,
                batch = %s,
                generation = %s,
                major_id = %s,
                teacher_id = %s,
                subject_id = %s,
                room_id = %s,
                years = %s,
                semester = %s,
                note = %s
            WHERE id = %s
        """, (study_sessions_id, group_student, batch, generation, major_id, teacher_id,
              subject_id, room_id, years, semester, note, timetable_id))

        connection.commit()
        return jsonify({'message': 'Timetable entry updated successfully'}), 200

    except mysql.connector.Error as err:
        connection.rollback()
        return jsonify({'error': f'An error occurred: {err}'}), 500

    finally:
        cursor.close()
        connection.close()
